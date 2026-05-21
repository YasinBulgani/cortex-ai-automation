"""
tm_routes.py — Test Management (Manuel Test Yönetimi) API
Prefix: /api/tm/

Endpoints:
  Projects   : /api/tm/projects
  Modules    : /api/tm/projects/<id>/modules
  Test Cases : /api/tm/modules/<id>/testcases
  Sprints    : /api/tm/projects/<id>/sprints
  Test Runs  : /api/tm/projects/<id>/runs
  Bugs       : /api/tm/projects/<id>/bugs
  Reports    : /api/tm/projects/<id>/report
"""

from flask import Blueprint, request, jsonify, session
from core.db import (
    # Projects
    create_project, get_projects, get_project, update_project, delete_project,
    # Modules
    create_module, get_modules, update_module, delete_module,
    # Test Cases
    create_test_case, get_test_cases, get_test_case, update_test_case,
    delete_test_case, add_test_case_step, delete_test_case_step, bulk_create_test_cases,
    # Sprints
    create_sprint, get_sprints, delete_sprint,
    # Runs
    create_manual_test_run, get_manual_test_runs,
    get_manual_test_run_results, update_run_result, close_manual_test_run,
    # Bugs
    create_bug, get_bugs, update_bug_status, delete_bug,
    # Reports
    get_project_report
)

tm_bp = Blueprint('tm', __name__)


def current_user_id():
    return session.get('user_id')


# ─────────────────────────────────────────────────────────────────────────────
# PROJECTS
# ─────────────────────────────────────────────────────────────────────────────

@tm_bp.route("/api/tm/projects", methods=["GET"])
def tm_list_projects():
    return jsonify(get_projects())


@tm_bp.route("/api/tm/projects", methods=["POST"])
def tm_create_project():
    data = request.json or {}
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"error": "Proje adı gerekli"}), 400
    pid = create_project(name, data.get("description", ""), current_user_id())
    return jsonify({"ok": True, "id": pid}), 201


@tm_bp.route("/api/tm/projects/<int:pid>", methods=["GET"])
def tm_get_project(pid):
    p = get_project(pid)
    if not p:
        return jsonify({"error": "Proje bulunamadı"}), 404
    return jsonify(p)


@tm_bp.route("/api/tm/projects/<int:pid>", methods=["PUT"])
def tm_update_project(pid):
    data = request.json or {}
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"error": "Proje adı gerekli"}), 400
    update_project(pid, name, data.get("description", ""))
    return jsonify({"ok": True})


@tm_bp.route("/api/tm/projects/<int:pid>", methods=["DELETE"])
def tm_delete_project(pid):
    delete_project(pid)
    return jsonify({"ok": True})


# ─────────────────────────────────────────────────────────────────────────────
# MODULES
# ─────────────────────────────────────────────────────────────────────────────

@tm_bp.route("/api/tm/projects/<int:pid>/modules", methods=["GET"])
def tm_list_modules(pid):
    return jsonify(get_modules(pid))


@tm_bp.route("/api/tm/projects/<int:pid>/modules", methods=["POST"])
def tm_create_module(pid):
    data = request.json or {}
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"error": "Modül adı gerekli"}), 400
    mid = create_module(pid, name, data.get("description", ""))
    return jsonify({"ok": True, "id": mid}), 201


@tm_bp.route("/api/tm/modules/<int:mid>", methods=["PUT"])
def tm_update_module(mid):
    data = request.json or {}
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"error": "Modül adı gerekli"}), 400
    update_module(mid, name, data.get("description", ""))
    return jsonify({"ok": True})


@tm_bp.route("/api/tm/modules/<int:mid>", methods=["DELETE"])
def tm_delete_module(mid):
    delete_module(mid)
    return jsonify({"ok": True})


# ─────────────────────────────────────────────────────────────────────────────
# TEST CASES
# ─────────────────────────────────────────────────────────────────────────────

@tm_bp.route("/api/tm/modules/<int:mid>/testcases", methods=["GET"])
def tm_list_testcases(mid):
    return jsonify(get_test_cases(mid))


@tm_bp.route("/api/tm/modules/<int:mid>/testcases", methods=["POST"])
def tm_create_testcase(mid):
    data = request.json or {}
    title = data.get("title", "").strip()
    if not title:
        return jsonify({"error": "Test case başlığı gerekli"}), 400
    tc_id = create_test_case(
        module_id=mid,
        title=title,
        description=data.get("description", ""),
        preconditions=data.get("preconditions", ""),
        priority=data.get("priority", "P2"),
        tags=data.get("tags", ""),
        created_by=current_user_id()
    )
    for step in data.get("steps", []):
        add_test_case_step(tc_id, step.get("action", ""), step.get("expected", ""))
    return jsonify({"ok": True, "id": tc_id}), 201


@tm_bp.route("/api/tm/testcases/<int:tc_id>", methods=["GET"])
def tm_get_testcase(tc_id):
    tc = get_test_case(tc_id)
    if not tc:
        return jsonify({"error": "Test case bulunamadı"}), 404
    return jsonify(tc)


@tm_bp.route("/api/tm/testcases/<int:tc_id>", methods=["PUT"])
def tm_update_testcase(tc_id):
    data = request.json or {}
    title = data.get("title", "").strip()
    if not title:
        return jsonify({"error": "Başlık gerekli"}), 400
    update_test_case(
        tc_id,
        title=title,
        description=data.get("description", ""),
        preconditions=data.get("preconditions", ""),
        priority=data.get("priority", "P2"),
        tags=data.get("tags", "")
    )
    return jsonify({"ok": True})


@tm_bp.route("/api/tm/testcases/<int:tc_id>", methods=["DELETE"])
def tm_delete_testcase(tc_id):
    delete_test_case(tc_id)
    return jsonify({"ok": True})


@tm_bp.route("/api/tm/testcases/<int:tc_id>/steps", methods=["POST"])
def tm_add_step(tc_id):
    data = request.json or {}
    action = data.get("action", "").strip()
    expected = data.get("expected", "").strip()
    if not action or not expected:
        return jsonify({"error": "Aksiyon ve beklenen sonuç gerekli"}), 400
    step_id = add_test_case_step(tc_id, action, expected)
    return jsonify({"ok": True, "id": step_id}), 201


@tm_bp.route("/api/tm/steps/<int:step_id>", methods=["DELETE"])
def tm_delete_step(step_id):
    delete_test_case_step(step_id)
    return jsonify({"ok": True})


@tm_bp.route("/api/tm/modules/<int:mid>/testcases/bulk", methods=["POST"])
def tm_bulk_create_testcases(mid):
    """AI önizlemesinden onaylanan test case'leri toplu kaydeder."""
    data = request.json or {}
    cases = data.get("cases", [])
    if not cases:
        return jsonify({"error": "Kaydedilecek test case yok"}), 400
    ids = bulk_create_test_cases(mid, cases, current_user_id())
    return jsonify({"ok": True, "created": len(ids), "ids": ids}), 201


# ─────────────────────────────────────────────────────────────────────────────
# SPRINTS
# ─────────────────────────────────────────────────────────────────────────────

@tm_bp.route("/api/tm/projects/<int:pid>/sprints", methods=["GET"])
def tm_list_sprints(pid):
    return jsonify(get_sprints(pid))


@tm_bp.route("/api/tm/projects/<int:pid>/sprints", methods=["POST"])
def tm_create_sprint(pid):
    data = request.json or {}
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"error": "Sprint adı gerekli"}), 400
    sid = create_sprint(
        pid, name,
        data.get("release_version", ""),
        data.get("start_date"),
        data.get("end_date")
    )
    return jsonify({"ok": True, "id": sid}), 201


@tm_bp.route("/api/tm/sprints/<int:sid>", methods=["DELETE"])
def tm_delete_sprint(sid):
    delete_sprint(sid)
    return jsonify({"ok": True})


# ─────────────────────────────────────────────────────────────────────────────
# TEST RUNS
# ─────────────────────────────────────────────────────────────────────────────

@tm_bp.route("/api/tm/projects/<int:pid>/runs", methods=["GET"])
def tm_list_runs(pid):
    return jsonify(get_manual_test_runs(pid))


@tm_bp.route("/api/tm/projects/<int:pid>/runs", methods=["POST"])
def tm_create_run(pid):
    data = request.json or {}
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"error": "Run adı gerekli"}), 400
    run_id = create_manual_test_run(
        project_id=pid,
        name=name,
        sprint_id=data.get("sprint_id"),
        started_by=current_user_id()
    )
    return jsonify({"ok": True, "id": run_id}), 201


@tm_bp.route("/api/tm/runs/<int:run_id>/results", methods=["GET"])
def tm_get_run_results(run_id):
    return jsonify(get_manual_test_run_results(run_id))


@tm_bp.route("/api/tm/results/<int:result_id>", methods=["PUT"])
def tm_update_result(result_id):
    data = request.json or {}
    status = data.get("status")
    if not status:
        return jsonify({"error": "Durum gerekli"}), 400
    update_run_result(
        result_id,
        status=status,
        actual_result=data.get("actual_result", ""),
        notes=data.get("notes", ""),
        executed_by=current_user_id()
    )
    return jsonify({"ok": True})


@tm_bp.route("/api/tm/runs/<int:run_id>/close", methods=["POST"])
def tm_close_run(run_id):
    close_manual_test_run(run_id)
    return jsonify({"ok": True})


# ─────────────────────────────────────────────────────────────────────────────
# BUGS
# ─────────────────────────────────────────────────────────────────────────────

@tm_bp.route("/api/tm/projects/<int:pid>/bugs", methods=["GET"])
def tm_list_bugs(pid):
    return jsonify(get_bugs(pid))


@tm_bp.route("/api/tm/bugs", methods=["POST"])
def tm_create_bug():
    data = request.json or {}
    title = data.get("title", "").strip()
    if not title:
        return jsonify({"error": "Bug başlığı gerekli"}), 400
    bug_id = create_bug(
        title=title,
        description=data.get("description", ""),
        severity=data.get("severity", "Medium"),
        result_id=data.get("result_id"),
        test_case_id=data.get("test_case_id"),
        created_by=current_user_id()
    )
    return jsonify({"ok": True, "id": bug_id}), 201


@tm_bp.route("/api/tm/bugs/<int:bug_id>", methods=["PUT"])
def tm_update_bug(bug_id):
    data = request.json or {}
    status = data.get("status")
    if not status:
        return jsonify({"error": "Durum gerekli"}), 400
    update_bug_status(bug_id, status)
    return jsonify({"ok": True})


@tm_bp.route("/api/tm/bugs/<int:bug_id>", methods=["DELETE"])
def tm_delete_bug(bug_id):
    delete_bug(bug_id)
    return jsonify({"ok": True})


# ─────────────────────────────────────────────────────────────────────────────
# REPORTS
# ─────────────────────────────────────────────────────────────────────────────

@tm_bp.route("/api/tm/projects/<int:pid>/report", methods=["GET"])
def tm_project_report(pid):
    return jsonify(get_project_report(pid))


@tm_bp.route("/api/tm/projects/<int:pid>/report/excel", methods=["GET"])
def tm_export_excel(pid):
    """Proje test case'lerini ve run sonuçlarını Excel olarak export eder."""
    try:
        import openpyxl
        from flask import send_file
        import io
        from core.db import get_modules, get_test_cases, get_manual_test_runs, get_bugs

        wb = openpyxl.Workbook()

        # ── Sheet 1: Test Cases ──
        ws_cases = wb.active
        ws_cases.title = "Test Cases"
        ws_cases.append(["Modül", "Başlık", "Öncelik", "Etiketler", "Açıklama", "Ön Koşul", "Adım Sayısı"])
        for mod in get_modules(pid):
            for tc in get_test_cases(mod['id']):
                ws_cases.append([mod['name'], tc['title'], tc['priority'], tc.get('tags',''), tc.get('description',''), tc.get('preconditions',''), len(tc.get('steps',[]))])

        # ── Sheet 2: Test Runs ──
        ws_runs = wb.create_sheet("Test Runs")
        ws_runs.append(["Run Adı", "Sprint", "Durum", "Pass", "Fail", "Not Run", "Tarih"])
        for run in get_manual_test_runs(pid):
            stats = run.get('stats', {})
            ws_runs.append([run['name'], run.get('sprint_name',''), run['status'], stats.get('Pass',0), stats.get('Fail',0), stats.get('Not Run',0), run['created_at']])

        # ── Sheet 3: Bugs ──
        ws_bugs = wb.create_sheet("Bugs")
        ws_bugs.append(["Başlık", "Önem", "Durum", "Jira Key", "Tarih"])
        for bug in get_bugs(pid):
            ws_bugs.append([bug['title'], bug['severity'], bug['status'], bug.get('jira_key',''), bug['created_at']])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=f'test_report_project_{pid}.xlsx')
    except ImportError:
        return jsonify({"error": "openpyxl kütüphanesi gerekli: pip install openpyxl"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500
