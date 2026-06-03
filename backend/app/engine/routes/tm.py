"""
Test Management (Manuel Test Yönetimi) routes — Flask engine'den FastAPI'ye port edilmiştir.

ÖNCE (Flask):
  /engine/routes/tm_routes.py — Blueprint, port 5001

SONRA (FastAPI):
  /backend/app/engine/routes/tm.py — APIRouter, port 8000 (consolidated)

Endpoints:
  Projects   : /api/tm/projects
  Modules    : /api/tm/projects/{pid}/modules
  Test Cases : /api/tm/modules/{mid}/testcases
  Sprints    : /api/tm/projects/{pid}/sprints
  Test Runs  : /api/tm/projects/{pid}/runs
  Bugs       : /api/tm/projects/{pid}/bugs
  Reports    : /api/tm/projects/{pid}/report
"""

import io
import logging
from typing import Optional

from fastapi import APIRouter, HTTPException, Response, status
from pydantic import BaseModel, Field, field_validator

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/tm", tags=["engine", "test-management"])


# ─── Schemas ─────────────────────────────────────────────────────────────────

class ProjectCreate(BaseModel):
    name: str = Field(..., min_length=1)
    description: str = ""


class ProjectUpdate(BaseModel):
    name: str = Field(..., min_length=1)
    description: str = ""


class ModuleCreate(BaseModel):
    name: str = Field(..., min_length=1)
    description: str = ""


class ModuleUpdate(BaseModel):
    name: str = Field(..., min_length=1)
    description: str = ""


class StepIn(BaseModel):
    action: str = ""
    expected: str = ""


class TestCaseCreate(BaseModel):
    title: str = Field(..., min_length=1)
    description: str = ""
    preconditions: str = ""
    priority: str = "P2"
    tags: str = ""
    steps: list[StepIn] = []

    @field_validator("priority")
    @classmethod
    def validate_priority(cls, v: str) -> str:
        allowed = {"P1", "P2", "P3", "P4"}
        if v not in allowed:
            raise ValueError(f"Geçersiz priority. İzin verilenler: {allowed}")
        return v


class TestCaseUpdate(BaseModel):
    title: str = Field(..., min_length=1)
    description: str = ""
    preconditions: str = ""
    priority: str = "P2"
    tags: str = ""

    @field_validator("priority")
    @classmethod
    def validate_priority(cls, v: str) -> str:
        allowed = {"P1", "P2", "P3", "P4"}
        if v not in allowed:
            raise ValueError(f"Geçersiz priority. İzin verilenler: {allowed}")
        return v


class StepCreate(BaseModel):
    action: str = Field(..., min_length=1)
    expected: str = Field(..., min_length=1)


class BulkCasesCreate(BaseModel):
    cases: list[TestCaseCreate] = []


class SprintCreate(BaseModel):
    name: str = Field(..., min_length=1)
    release_version: str = ""
    start_date: Optional[str] = None
    end_date: Optional[str] = None


class RunCreate(BaseModel):
    name: str = Field(..., min_length=1)
    sprint_id: Optional[int] = None


_VALID_RESULT_STATUSES = {"Pass", "Fail", "Blocked", "Not Run", "Skipped"}
_VALID_BUG_STATUSES = {"Open", "In Progress", "Resolved", "Closed", "Reopened"}
_VALID_SEVERITIES = {"Critical", "High", "Medium", "Low"}
_VALID_PRIORITIES = {"P1", "P2", "P3", "P4"}


class ResultUpdate(BaseModel):
    status: str = Field(..., min_length=1)
    actual_result: str = ""
    notes: str = ""

    @field_validator("status")
    @classmethod
    def validate_status(cls, v: str) -> str:
        if v not in _VALID_RESULT_STATUSES:
            raise ValueError(f"Geçersiz status. İzin verilenler: {_VALID_RESULT_STATUSES}")
        return v


class BugCreate(BaseModel):
    title: str = Field(..., min_length=1)
    description: str = ""
    severity: str = "Medium"
    result_id: Optional[int] = None
    test_case_id: Optional[int] = None

    @field_validator("severity")
    @classmethod
    def validate_severity(cls, v: str) -> str:
        if v not in _VALID_SEVERITIES:
            raise ValueError(f"Geçersiz severity. İzin verilenler: {_VALID_SEVERITIES}")
        return v


class BugStatusUpdate(BaseModel):
    status: str = Field(..., min_length=1)

    @field_validator("status")
    @classmethod
    def validate_status(cls, v: str) -> str:
        if v not in _VALID_BUG_STATUSES:
            raise ValueError(f"Geçersiz status. İzin verilenler: {_VALID_BUG_STATUSES}")
        return v


# ─── In-memory store (placeholder) ───────────────────────────────────────────

class _TmStore:
    """Geçici in-memory store. Migration tamamlanınca SQLAlchemy repository kullanılacak."""

    def __init__(self):
        self._projects: dict[int, dict] = {}
        self._modules: dict[int, dict] = {}
        self._test_cases: dict[int, dict] = {}
        self._steps: dict[int, dict] = {}
        self._sprints: dict[int, dict] = {}
        self._runs: dict[int, dict] = {}
        self._results: dict[int, dict] = {}
        self._bugs: dict[int, dict] = {}
        self._next: dict[str, int] = {k: 1 for k in ("p", "m", "tc", "s", "sp", "r", "res", "b")}

    def _id(self, key: str) -> int:
        v = self._next[key]
        self._next[key] += 1
        return v

    # ── Projects ──────────────────────────────────────────────────────────────

    def create_project(self, name: str, description: str, user_id: Optional[int]) -> int:
        pid = self._id("p")
        self._projects[pid] = {"id": pid, "name": name, "description": description, "created_by": user_id}
        return pid

    def get_projects(self) -> list[dict]:
        return list(self._projects.values())

    def get_project(self, pid: int) -> Optional[dict]:
        return self._projects.get(pid)

    def update_project(self, pid: int, name: str, description: str) -> None:
        if pid in self._projects:
            self._projects[pid].update({"name": name, "description": description})

    def delete_project(self, pid: int) -> None:
        self._projects.pop(pid, None)

    # ── Modules ───────────────────────────────────────────────────────────────

    def create_module(self, pid: int, name: str, description: str) -> int:
        mid = self._id("m")
        self._modules[mid] = {"id": mid, "project_id": pid, "name": name, "description": description}
        return mid

    def get_modules(self, pid: int) -> list[dict]:
        return [m for m in self._modules.values() if m["project_id"] == pid]

    def update_module(self, mid: int, name: str, description: str) -> None:
        if mid in self._modules:
            self._modules[mid].update({"name": name, "description": description})

    def delete_module(self, mid: int) -> None:
        self._modules.pop(mid, None)

    # ── Test Cases ────────────────────────────────────────────────────────────

    def create_test_case(self, mid: int, title: str, description: str, preconditions: str,
                         priority: str, tags: str, user_id: Optional[int]) -> int:
        tc_id = self._id("tc")
        self._test_cases[tc_id] = {
            "id": tc_id, "module_id": mid, "title": title, "description": description,
            "preconditions": preconditions, "priority": priority, "tags": tags,
            "created_by": user_id, "steps": [],
        }
        return tc_id

    def get_test_cases(self, mid: int) -> list[dict]:
        return [tc for tc in self._test_cases.values() if tc["module_id"] == mid]

    def get_test_case(self, tc_id: int) -> Optional[dict]:
        return self._test_cases.get(tc_id)

    def update_test_case(self, tc_id: int, **kwargs) -> None:
        if tc_id in self._test_cases:
            self._test_cases[tc_id].update(kwargs)

    def delete_test_case(self, tc_id: int) -> None:
        self._test_cases.pop(tc_id, None)

    def add_step(self, tc_id: int, action: str, expected: str) -> int:
        step_id = self._id("s")
        self._steps[step_id] = {"id": step_id, "tc_id": tc_id, "action": action, "expected": expected}
        if tc_id in self._test_cases:
            self._test_cases[tc_id].setdefault("steps", []).append(self._steps[step_id])
        return step_id

    def delete_step(self, step_id: int) -> None:
        step = self._steps.pop(step_id, None)
        if step:
            tc = self._test_cases.get(step["tc_id"])
            if tc:
                tc["steps"] = [s for s in tc.get("steps", []) if s["id"] != step_id]

    def bulk_create_test_cases(self, mid: int, cases: list[dict], user_id: Optional[int]) -> list[int]:
        ids = []
        for c in cases:
            tc_id = self.create_test_case(
                mid=mid,
                title=c.get("title", ""),
                description=c.get("description", ""),
                preconditions=c.get("preconditions", ""),
                priority=c.get("priority", "P2"),
                tags=c.get("tags", ""),
                user_id=user_id,
            )
            for step in c.get("steps", []):
                self.add_step(tc_id, step.get("action", ""), step.get("expected", ""))
            ids.append(tc_id)
        return ids

    # ── Sprints ───────────────────────────────────────────────────────────────

    def create_sprint(self, pid: int, name: str, release_version: str,
                      start_date: Optional[str], end_date: Optional[str]) -> int:
        sid = self._id("sp")
        self._sprints[sid] = {
            "id": sid, "project_id": pid, "name": name,
            "release_version": release_version, "start_date": start_date, "end_date": end_date,
        }
        return sid

    def get_sprints(self, pid: int) -> list[dict]:
        return [s for s in self._sprints.values() if s["project_id"] == pid]

    def delete_sprint(self, sid: int) -> None:
        self._sprints.pop(sid, None)

    # ── Runs ──────────────────────────────────────────────────────────────────

    def create_run(self, pid: int, name: str, sprint_id: Optional[int], user_id: Optional[int]) -> int:
        run_id = self._id("r")
        self._runs[run_id] = {
            "id": run_id, "project_id": pid, "name": name,
            "sprint_id": sprint_id, "started_by": user_id, "status": "open",
        }
        return run_id

    def get_runs(self, pid: int) -> list[dict]:
        return [r for r in self._runs.values() if r["project_id"] == pid]

    def get_run_results(self, run_id: int) -> list[dict]:
        return [r for r in self._results.values() if r["run_id"] == run_id]

    def update_result(self, result_id: int, run_status: str, actual_result: str,
                      notes: str, user_id: Optional[int]) -> None:
        if result_id in self._results:
            self._results[result_id].update({
                "status": run_status, "actual_result": actual_result,
                "notes": notes, "executed_by": user_id,
            })

    def close_run(self, run_id: int) -> None:
        if run_id in self._runs:
            self._runs[run_id]["status"] = "closed"

    # ── Bugs ──────────────────────────────────────────────────────────────────

    def create_bug(self, title: str, description: str, severity: str,
                   result_id: Optional[int], test_case_id: Optional[int],
                   user_id: Optional[int]) -> int:
        bug_id = self._id("b")
        self._bugs[bug_id] = {
            "id": bug_id, "title": title, "description": description, "severity": severity,
            "result_id": result_id, "test_case_id": test_case_id,
            "created_by": user_id, "status": "Open",
        }
        return bug_id

    def get_bugs(self, pid: int) -> list[dict]:
        # In-memory store has no project linkage — return all as placeholder
        return list(self._bugs.values())

    def update_bug_status(self, bug_id: int, new_status: str) -> None:
        if bug_id in self._bugs:
            self._bugs[bug_id]["status"] = new_status

    def delete_bug(self, bug_id: int) -> None:
        self._bugs.pop(bug_id, None)

    # ── Reports ───────────────────────────────────────────────────────────────

    def get_project_report(self, pid: int) -> dict:
        modules = self.get_modules(pid)
        total_cases = sum(len(self.get_test_cases(m["id"])) for m in modules)
        runs = self.get_runs(pid)
        bugs = self.get_bugs(pid)
        return {
            "project_id": pid,
            "module_count": len(modules),
            "test_case_count": total_cases,
            "run_count": len(runs),
            "bug_count": len(bugs),
        }


_store = _TmStore()


# ─── Routes: Projects ────────────────────────────────────────────────────────

@router.get("/projects", response_model=list[dict])
def tm_list_projects() -> list[dict]:
    """Tüm projeleri listeler."""
    return _store.get_projects()


@router.post("/projects", status_code=status.HTTP_201_CREATED, response_model=dict)
def tm_create_project(body: ProjectCreate) -> dict:
    """Yeni bir proje oluşturur."""
    pid = _store.create_project(body.name.strip(), body.description, user_id=None)
    return {"ok": True, "id": pid}


@router.get("/projects/{pid}", response_model=dict)
def tm_get_project(pid: int) -> dict:
    """Belirli bir projeyi id'ye göre getirir."""
    p = _store.get_project(pid)
    if not p:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Proje bulunamadı")
    return p


@router.put("/projects/{pid}", response_model=dict)
def tm_update_project(pid: int, body: ProjectUpdate) -> dict:
    """Bir projenin adını ve açıklamasını günceller."""
    if not _store.get_project(pid):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Proje bulunamadı")
    _store.update_project(pid, body.name.strip(), body.description)
    return {"ok": True}


@router.delete("/projects/{pid}", status_code=status.HTTP_200_OK, response_model=dict)
def tm_delete_project(pid: int) -> dict:
    """Bir projeyi siler."""
    if not _store.get_project(pid):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Proje bulunamadı")
    _store.delete_project(pid)
    return {"ok": True}


# ─── Routes: Modules ─────────────────────────────────────────────────────────

@router.get("/projects/{pid}/modules", response_model=list[dict])
def tm_list_modules(pid: int) -> list[dict]:
    """Bir projenin tüm modüllerini listeler."""
    if not _store.get_project(pid):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Proje bulunamadı")
    return _store.get_modules(pid)


@router.post("/projects/{pid}/modules", status_code=status.HTTP_201_CREATED, response_model=dict)
def tm_create_module(pid: int, body: ModuleCreate) -> dict:
    """Bir proje altında yeni modül oluşturur."""
    if not _store.get_project(pid):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Proje bulunamadı")
    mid = _store.create_module(pid, body.name.strip(), body.description)
    return {"ok": True, "id": mid}


@router.put("/modules/{mid}", response_model=dict)
def tm_update_module(mid: int, body: ModuleUpdate) -> dict:
    """Bir modülün adını ve açıklamasını günceller."""
    if mid not in _store._modules:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Modül bulunamadı")
    _store.update_module(mid, body.name.strip(), body.description)
    return {"ok": True}


@router.delete("/modules/{mid}", status_code=status.HTTP_200_OK, response_model=dict)
def tm_delete_module(mid: int) -> dict:
    """Bir modülü siler."""
    if mid not in _store._modules:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Modül bulunamadı")
    _store.delete_module(mid)
    return {"ok": True}


# ─── Routes: Test Cases ──────────────────────────────────────────────────────

@router.get("/modules/{mid}/testcases", response_model=list[dict])
def tm_list_testcases(mid: int) -> list[dict]:
    """Bir modüle ait tüm test case'leri listeler."""
    if mid not in _store._modules:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Modül bulunamadı")
    return _store.get_test_cases(mid)


@router.post("/modules/{mid}/testcases", status_code=status.HTTP_201_CREATED, response_model=dict)
def tm_create_testcase(mid: int, body: TestCaseCreate) -> dict:
    """Bir modül altında yeni test case oluşturur; adımları birlikte kaydeder."""
    if mid not in _store._modules:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Modül bulunamadı")
    tc_id = _store.create_test_case(
        mid=mid,
        title=body.title.strip(),
        description=body.description,
        preconditions=body.preconditions,
        priority=body.priority,
        tags=body.tags,
        user_id=None,
    )
    for step in body.steps:
        _store.add_step(tc_id, step.action, step.expected)
    return {"ok": True, "id": tc_id}


@router.get("/testcases/{tc_id}", response_model=dict)
def tm_get_testcase(tc_id: int) -> dict:
    """Belirli bir test case'i adımlarıyla birlikte getirir."""
    tc = _store.get_test_case(tc_id)
    if not tc:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Test case bulunamadı")
    return tc


@router.put("/testcases/{tc_id}", response_model=dict)
def tm_update_testcase(tc_id: int, body: TestCaseUpdate) -> dict:
    """Bir test case'in meta verilerini günceller (adımlar dahil değil)."""
    if not _store.get_test_case(tc_id):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Test case bulunamadı")
    _store.update_test_case(
        tc_id,
        title=body.title.strip(),
        description=body.description,
        preconditions=body.preconditions,
        priority=body.priority,
        tags=body.tags,
    )
    return {"ok": True}


@router.delete("/testcases/{tc_id}", status_code=status.HTTP_200_OK, response_model=dict)
def tm_delete_testcase(tc_id: int) -> dict:
    """Bir test case'i ve bağlı adımlarını siler."""
    if not _store.get_test_case(tc_id):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Test case bulunamadı")
    _store.delete_test_case(tc_id)
    return {"ok": True}


@router.post("/testcases/{tc_id}/steps", status_code=status.HTTP_201_CREATED, response_model=dict)
def tm_add_step(tc_id: int, body: StepCreate) -> dict:
    """Bir test case'e yeni adım ekler."""
    if not _store.get_test_case(tc_id):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Test case bulunamadı")
    step_id = _store.add_step(tc_id, body.action.strip(), body.expected.strip())
    return {"ok": True, "id": step_id}


@router.delete("/steps/{step_id}", status_code=status.HTTP_200_OK, response_model=dict)
def tm_delete_step(step_id: int) -> dict:
    """Bir test adımını siler."""
    if step_id not in _store._steps:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Adım bulunamadı")
    _store.delete_step(step_id)
    return {"ok": True}


@router.post("/modules/{mid}/testcases/bulk", status_code=status.HTTP_201_CREATED, response_model=dict)
def tm_bulk_create_testcases(mid: int, body: BulkCasesCreate) -> dict:
    """AI önizlemesinden onaylanan test case'leri toplu kaydeder."""
    if mid not in _store._modules:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Modül bulunamadı")
    if not body.cases:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Kaydedilecek test case yok")
    ids = _store.bulk_create_test_cases(mid, [c.model_dump() for c in body.cases], user_id=None)
    return {"ok": True, "created": len(ids), "ids": ids}


# ─── Routes: Sprints ─────────────────────────────────────────────────────────

@router.get("/projects/{pid}/sprints", response_model=list[dict])
def tm_list_sprints(pid: int) -> list[dict]:
    """Bir projenin tüm sprint'lerini listeler."""
    if not _store.get_project(pid):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Proje bulunamadı")
    return _store.get_sprints(pid)


@router.post("/projects/{pid}/sprints", status_code=status.HTTP_201_CREATED, response_model=dict)
def tm_create_sprint(pid: int, body: SprintCreate) -> dict:
    """Bir proje için yeni sprint oluşturur."""
    if not _store.get_project(pid):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Proje bulunamadı")
    sid = _store.create_sprint(
        pid, body.name.strip(), body.release_version, body.start_date, body.end_date
    )
    return {"ok": True, "id": sid}


@router.delete("/sprints/{sid}", status_code=status.HTTP_200_OK, response_model=dict)
def tm_delete_sprint(sid: int) -> dict:
    """Bir sprint'i siler."""
    if sid not in _store._sprints:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Sprint bulunamadı")
    _store.delete_sprint(sid)
    return {"ok": True}


# ─── Routes: Test Runs ───────────────────────────────────────────────────────

@router.get("/projects/{pid}/runs", response_model=list[dict])
def tm_list_runs(pid: int) -> list[dict]:
    """Bir projenin tüm test run'larını listeler."""
    if not _store.get_project(pid):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Proje bulunamadı")
    return _store.get_runs(pid)


@router.post("/projects/{pid}/runs", status_code=status.HTTP_201_CREATED, response_model=dict)
def tm_create_run(pid: int, body: RunCreate) -> dict:
    """Bir proje için yeni test run başlatır."""
    if not _store.get_project(pid):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Proje bulunamadı")
    run_id = _store.create_run(pid, body.name.strip(), body.sprint_id, user_id=None)
    return {"ok": True, "id": run_id}


@router.get("/runs/{run_id}/results", response_model=list[dict])
def tm_get_run_results(run_id: int) -> list[dict]:
    """Bir test run'ın tüm sonuçlarını listeler."""
    if run_id not in _store._runs:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Test run bulunamadı")
    return _store.get_run_results(run_id)


@router.put("/results/{result_id}", response_model=dict)
def tm_update_result(result_id: int, body: ResultUpdate) -> dict:
    """Bir test sonucunun durumunu ve notlarını günceller."""
    if result_id not in _store._results:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Sonuç bulunamadı")
    _store.update_result(result_id, body.status, body.actual_result, body.notes, user_id=None)
    return {"ok": True}


@router.post("/runs/{run_id}/close", response_model=dict)
def tm_close_run(run_id: int) -> dict:
    """Açık bir test run'ı kapatır."""
    if run_id not in _store._runs:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Test run bulunamadı")
    _store.close_run(run_id)
    return {"ok": True}


# ─── Routes: Bugs ────────────────────────────────────────────────────────────

@router.get("/projects/{pid}/bugs", response_model=list[dict])
def tm_list_bugs(pid: int) -> list[dict]:
    """Bir projeye ait tüm bug'ları listeler."""
    if not _store.get_project(pid):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Proje bulunamadı")
    return _store.get_bugs(pid)


@router.post("/bugs", status_code=status.HTTP_201_CREATED, response_model=dict)
def tm_create_bug(body: BugCreate) -> dict:
    """Yeni bir bug kaydı oluşturur."""
    bug_id = _store.create_bug(
        title=body.title.strip(),
        description=body.description,
        severity=body.severity,
        result_id=body.result_id,
        test_case_id=body.test_case_id,
        user_id=None,
    )
    return {"ok": True, "id": bug_id}


@router.put("/bugs/{bug_id}", response_model=dict)
def tm_update_bug(bug_id: int, body: BugStatusUpdate) -> dict:
    """Bir bug'ın durumunu günceller."""
    if bug_id not in _store._bugs:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Bug bulunamadı")
    _store.update_bug_status(bug_id, body.status)
    return {"ok": True}


@router.delete("/bugs/{bug_id}", status_code=status.HTTP_200_OK, response_model=dict)
def tm_delete_bug(bug_id: int) -> dict:
    """Bir bug kaydını siler."""
    if bug_id not in _store._bugs:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Bug bulunamadı")
    _store.delete_bug(bug_id)
    return {"ok": True}


# ─── Routes: Reports ─────────────────────────────────────────────────────────

@router.get("/projects/{pid}/report", response_model=dict)
def tm_project_report(pid: int) -> dict:
    """Bir projenin modül, test case, run ve bug sayılarını özetleyen rapor döner."""
    if not _store.get_project(pid):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Proje bulunamadı")
    return _store.get_project_report(pid)


@router.get("/projects/{pid}/report/excel")
def tm_export_excel(pid: int) -> Response:
    """Proje test case'lerini ve run sonuçlarını Excel olarak export eder."""
    if not _store.get_project(pid):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Proje bulunamadı")

    try:
        import openpyxl  # type: ignore[import]
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="openpyxl kütüphanesi gerekli: pip install openpyxl",
        )

    try:
        wb = openpyxl.Workbook()

        # Sheet 1: Test Cases
        ws_cases = wb.active
        ws_cases.title = "Test Cases"
        ws_cases.append(["Modül", "Başlık", "Öncelik", "Etiketler", "Açıklama", "Ön Koşul", "Adım Sayısı"])
        for mod in _store.get_modules(pid):
            for tc in _store.get_test_cases(mod["id"]):
                ws_cases.append([
                    mod["name"], tc["title"], tc["priority"],
                    tc.get("tags", ""), tc.get("description", ""),
                    tc.get("preconditions", ""), len(tc.get("steps", [])),
                ])

        # Sheet 2: Test Runs
        ws_runs = wb.create_sheet("Test Runs")
        ws_runs.append(["Run Adı", "Sprint", "Durum", "Pass", "Fail", "Not Run", "Tarih"])
        for run in _store.get_runs(pid):
            stats = run.get("stats", {})
            ws_runs.append([
                run["name"], run.get("sprint_name", ""), run["status"],
                stats.get("Pass", 0), stats.get("Fail", 0), stats.get("Not Run", 0),
                run.get("created_at", ""),
            ])

        # Sheet 3: Bugs
        ws_bugs = wb.create_sheet("Bugs")
        ws_bugs.append(["Başlık", "Önem", "Durum", "Jira Key", "Tarih"])
        for bug in _store.get_bugs(pid):
            ws_bugs.append([
                bug["title"], bug["severity"], bug["status"],
                bug.get("jira_key", ""), bug.get("created_at", ""),
            ])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        return Response(
            content=buf.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=test_report_project_{pid}.xlsx"},
        )
    except Exception as exc:
        logger.error("Excel export failed for project %s: %s", pid, exc)
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Excel oluşturma hatası")
