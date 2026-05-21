"""Excel artifact generation for durable AI workflows."""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from app.config import settings


HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def build_workflow_excel_report(
    workflow_id: str,
    state: dict[str, Any],
    *,
    events: list[dict[str, Any]] | None = None,
    artifacts: list[dict[str, Any]] | None = None,
) -> str:
    """Create a standard workbook for a workflow run and return its path."""
    out_dir = Path(settings.artifacts_dir) / "agents_v2" / workflow_id
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "run_report.xlsx"

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"

    events = events or []
    artifacts = artifacts or []

    _write_key_values(
        summary,
        [
            ("workflow_id", workflow_id),
            ("status", state.get("status")),
            ("project_id", state.get("project_id")),
            ("input_source", state.get("input_source")),
            ("workflow_type", state.get("workflow_type")),
            ("dry_run", state.get("dry_run")),
            ("requires_approval", state.get("requires_approval")),
            ("created_at", state.get("created_at")),
            ("completed_at", state.get("completed_at")),
            ("tokens_used", state.get("tokens_used", 0)),
            ("llm_calls_count", state.get("llm_calls_count", 0)),
            ("cost_usd", state.get("cost_usd", 0.0)),
            ("error", state.get("error") or state.get("error_message")),
        ],
    )

    _write_rows(
        wb.create_sheet("Scenarios"),
        ["name", "scenario_count", "feature_path", "risk_level", "raw"],
        [
            [
                scenario.get("name"),
                scenario.get("scenario_count"),
                scenario.get("feature_path"),
                scenario.get("risk_level"),
                _json(scenario),
            ]
            for scenario in _dicts(state.get("scenarios"))
        ],
    )

    _write_rows(
        wb.create_sheet("Generated Artifacts"),
        ["kind", "name", "storage_path", "mime_type", "size_bytes", "sha256", "created_at"],
        [
            [
                artifact.get("kind"),
                artifact.get("name"),
                artifact.get("storage_path"),
                artifact.get("mime_type"),
                artifact.get("size_bytes"),
                (artifact.get("metadata") or {}).get("sha256")
                if isinstance(artifact.get("metadata"), dict)
                else None,
                artifact.get("created_at"),
            ]
            for artifact in artifacts
        ],
    )

    review = state.get("review") if isinstance(state.get("review"), dict) else {}
    run_result = state.get("run_result") if isinstance(state.get("run_result"), dict) else {}
    _write_key_values(
        wb.create_sheet("Validation"),
        [
            ("passed_count", run_result.get("passed_count")),
            ("failed_count", run_result.get("failed_count")),
            ("code_quality_score", review.get("code_quality_score")),
            ("recommended_action", review.get("recommended_action")),
            ("errors", _json(state.get("errors", []))),
            ("healing_result", _json(state.get("healing_result"))),
        ],
    )

    _write_rows(
        wb.create_sheet("LLM Trace"),
        ["timestamp", "event_type", "agent_name", "message", "data"],
        [
            [
                event.get("timestamp"),
                event.get("event_type"),
                event.get("agent_name"),
                event.get("message"),
                _json(event.get("data", {})),
            ]
            for event in events
            if event.get("event_type") in {"llm_call", "agent_started", "agent_finished", "error"}
        ],
    )

    _write_key_values(
        wb.create_sheet("Cost"),
        [
            ("tokens_used", state.get("tokens_used", 0)),
            ("llm_calls_count", state.get("llm_calls_count", 0)),
            ("cost_usd", state.get("cost_usd", 0.0)),
            ("event_count", len(events)),
            ("artifact_count", len(artifacts)),
        ],
    )

    _write_rows(
        wb.create_sheet("Approvals"),
        ["timestamp", "decision", "actor_id", "note", "approval_id"],
        [
            [
                approval.get("created_at") or event.get("timestamp"),
                approval.get("decision"),
                approval.get("actor_id"),
                approval.get("note"),
                approval.get("approval_id"),
            ]
            for event in events
            if event.get("event_type") == "approval_recorded"
            for approval in [event.get("data") if isinstance(event.get("data"), dict) else {}]
        ],
    )

    for sheet in wb.worksheets:
        _autosize(sheet)

    wb.save(out_path)
    return str(out_path)


def _write_key_values(ws: Worksheet, rows: list[tuple[str, Any]]) -> None:
    _write_rows(ws, ["field", "value"], rows)


def _write_rows(ws: Worksheet, headers: list[str], rows: list[list[Any] | tuple[Any, ...]]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row in rows:
        ws.append([_cell(value) for value in row])


def _autosize(ws: Worksheet) -> None:
    for column in ws.columns:
        max_len = 0
        letter = column[0].column_letter
        for cell in column:
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 80)


def _cell(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        return _json(value)
    return value


def _dicts(value: Any) -> list[dict[str, Any]]:
    if not isinstance(value, list):
        return []
    return [item for item in value if isinstance(item, dict)]


def _json(value: Any) -> str:
    if value is None:
        return ""
    try:
        return json.dumps(value, ensure_ascii=False, default=str)
    except Exception:
        return str(value)
