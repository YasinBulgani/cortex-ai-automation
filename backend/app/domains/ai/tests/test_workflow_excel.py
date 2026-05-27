from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from app.config import settings
from app.domains.ai.workflow_excel import build_workflow_excel_report


def test_build_workflow_excel_report(tmp_path, monkeypatch):
    monkeypatch.setattr(settings, "artifacts_dir", str(tmp_path))

    path = build_workflow_excel_report(
        "run-1",
        {
            "status": "completed",
            "project_id": "p1",
            "input_source": "text",
            "workflow_type": "test_generation",
            "tokens_used": 120,
            "cost_usd": 0.12,
            "llm_calls_count": 3,
            "scenarios": [{"name": "Login", "scenario_count": 2}],
            "run_result": {"passed_count": 2, "failed_count": 0},
            "review": {"code_quality_score": 0.92, "recommended_action": "approve"},
        },
        events=[
            {"event_type": "llm_call", "timestamp": "2026-05-17T00:00:00", "data": {"model": "qwen"}},
            {
                "event_type": "approval_recorded",
                "timestamp": "2026-05-17T00:01:00",
                "data": {"decision": "approved", "actor_id": "u1", "approval_id": "a1"},
            },
        ],
        artifacts=[
            {
                "kind": "junit_xml",
                "name": "junit.xml",
                "storage_path": "/tmp/junit.xml",  # nosec B108
                "mime_type": "application/xml",
                "size_bytes": 42,
                "metadata": {"sha256": "abc123"},
            }
        ],
    )

    assert Path(path).exists()
    wb = load_workbook(path)
    assert wb.sheetnames == [
        "Summary",
        "Scenarios",
        "Generated Artifacts",
        "Validation",
        "LLM Trace",
        "Cost",
        "Approvals",
    ]
    assert wb["Summary"]["B2"].value == "run-1"
    assert wb["Scenarios"]["A2"].value == "Login"
    assert wb["Generated Artifacts"]["F1"].value == "sha256"
    assert wb["Generated Artifacts"]["F2"].value == "abc123"
