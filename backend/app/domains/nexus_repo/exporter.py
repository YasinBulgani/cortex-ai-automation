"""Nexus Repo — Dışa Aktarım Motoru.

Desteklenen formatlar:
  gherkin  → .feature dosyası (BDD/Cucumber/Pytest-BDD)
  postman  → Postman Collection v2.1 JSON
  excel    → .xlsx (openpyxl)
  jira     → Jira CSV (Test Management uyumlu — Xray/Zephyr import şeması)
"""

from __future__ import annotations

import csv
import io
import json
import logging
import os
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from sqlalchemy.orm import Session

from app.config import settings
from app.infra.database import SessionLocal
from .models import NexusCase, NexusExport, NexusProject, NexusScenario

_log = logging.getLogger(__name__)

_EXPORTS_BASE = Path(settings.artifacts_dir) / "nexus_exports"


def _ensure_dir(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


# ── Gherkin ───────────────────────────────────────────────────────────────────

def _to_gherkin(project: NexusProject, scenarios: list[NexusScenario]) -> str:
    """Senaryoları tek .feature dosyasına dönüştür."""
    lines: list[str] = [
        f"# Nexus Repo — {project.name}",
        f"# Üretildi: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
        f"# Repo: {project.repo_url}",
        "",
    ]

    feature_areas: dict[str, list[NexusScenario]] = {}
    for s in scenarios:
        area = s.feature_area or "Genel"
        feature_areas.setdefault(area, []).append(s)

    for area, area_scenarios in feature_areas.items():
        lines.append(f"Feature: {area}")
        lines.append("")
        for sc in area_scenarios:
            if sc.gherkin:
                # Hazır Gherkin varsa doğrudan ekle (Feature: satırını atla)
                gherkin_lines = sc.gherkin.strip().split("\n")
                # Feature: satırını atla
                body = [l for l in gherkin_lines if not l.strip().startswith("Feature:")]
                lines.extend(body)
            else:
                # Gherkin yoksa basit şablon üret
                lines.append(f"  Scenario: {sc.title}")
                lines.append(f"    # tip: {sc.type} | öncelik: {sc.priority}")
                if sc.notes:
                    for note_line in sc.notes.split("\n")[:3]:
                        lines.append(f"    # {note_line}")
                lines.append("    Given gerekli ön koşullar karşılanmış")
                lines.append("    When işlem gerçekleştirilir")
                lines.append("    Then beklenen sonuç doğrulanır")
            lines.append("")

    return "\n".join(lines)


# ── Postman Collection ────────────────────────────────────────────────────────

def _to_postman(project: NexusProject, scenarios: list[NexusScenario]) -> dict:
    """Postman Collection v2.1 JSON oluştur — servis tipi senaryolar için."""
    items: list[dict] = []

    service_scenarios = [s for s in scenarios if s.type == "service"]

    for sc in service_scenarios:
        # Gherkin'den endpoint ipuçlarını çıkarmaya çalış
        path = "/api/endpoint"
        method = "GET"
        if sc.gherkin:
            import re
            # "When I send a GET request to '/path'" gibi desenleri ara
            m = re.search(r"(GET|POST|PUT|PATCH|DELETE)\s+['\"]?([/\w{}-]+)['\"]?", sc.gherkin, re.IGNORECASE)
            if m:
                method = m.group(1).upper()
                path = m.group(2)

        item: dict = {
            "name": sc.title,
            "request": {
                "method": method,
                "header": [
                    {"key": "Content-Type", "value": "application/json"},
                    {"key": "Authorization", "value": "Bearer {{access_token}}"},
                ],
                "url": {
                    "raw": f"{{{{base_url}}}}{path}",
                    "host": ["{{base_url}}"],
                    "path": [p for p in path.split("/") if p],
                },
                "description": sc.notes or "",
            },
            "response": [],
        }

        # Test script — status code kontrolü
        item["event"] = [
            {
                "listen": "test",
                "script": {
                    "exec": [
                        "pm.test('Status code 2xx', function () {",
                        "    pm.expect(pm.response.code).to.be.oneOf([200, 201, 204]);",
                        "});",
                    ],
                    "type": "text/javascript",
                },
            }
        ]
        items.append(item)

    return {
        "info": {
            "_postman_id": str(uuid.uuid4()),
            "name": f"Nexus Repo — {project.name}",
            "description": f"Repo: {project.repo_url}\nÜretildi: {datetime.now(timezone.utc).isoformat()}",
            "schema": "https://schema.getpostman.com/json/collection/v2.1.0/collection.json",
        },
        "variable": [
            {"key": "base_url", "value": "http://localhost:8000", "type": "string"},
            {"key": "access_token", "value": "", "type": "string"},
        ],
        "item": items,
    }


# ── Excel ─────────────────────────────────────────────────────────────────────

def _to_excel(project: NexusProject, scenarios: list[NexusScenario], cases_by_scenario: dict[str, list[NexusCase]], out_path: Path) -> None:
    """openpyxl ile Excel raporu oluştur."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        _log.warning("openpyxl yüklü değil — Excel export atlandı")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Senaryoları"

    # Başlık satırı
    headers = ["#", "Başlık", "Tip", "Alan", "Öncelik", "Durum", "Test Case", "Adımlar", "Beklenen Sonuç", "Gherkin"]
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 24

    row = 2
    for sc_idx, sc in enumerate(scenarios, start=1):
        sc_cases = cases_by_scenario.get(sc.id, [])
        if not sc_cases:
            # Senaryo satırı (case yok)
            ws.cell(row=row, column=1, value=sc_idx)
            ws.cell(row=row, column=2, value=sc.title)
            ws.cell(row=row, column=3, value=sc.type)
            ws.cell(row=row, column=4, value=sc.feature_area or "")
            ws.cell(row=row, column=5, value=sc.priority)
            ws.cell(row=row, column=6, value=sc.status)
            ws.cell(row=row, column=10, value=sc.gherkin or "")
            row += 1
        else:
            for c_idx, case in enumerate(sc_cases):
                ws.cell(row=row, column=1, value=sc_idx)
                ws.cell(row=row, column=2, value=sc.title)
                ws.cell(row=row, column=3, value=sc.type)
                ws.cell(row=row, column=4, value=sc.feature_area or "")
                ws.cell(row=row, column=5, value=sc.priority)
                ws.cell(row=row, column=6, value=sc.status)
                ws.cell(row=row, column=7, value=case.name)
                # Adımlar
                steps_text = ""
                if case.steps:
                    steps_text = "\n".join(
                        f"{s.get('type','').upper()}: {s.get('step','')}" for s in case.steps if isinstance(s, dict)
                    )
                ws.cell(row=row, column=8, value=steps_text)
                ws.cell(row=row, column=9, value=case.expected_result or "")
                if c_idx == 0:
                    ws.cell(row=row, column=10, value=sc.gherkin or "")
                for col in range(1, 11):
                    ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")
                row += 1

    # Sütun genişlikleri
    col_widths = [5, 40, 12, 20, 12, 12, 35, 50, 35, 60]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # Özet sayfası
    ws2 = wb.create_sheet("Özet")
    ws2.append(["Nexus Repo — Test Senaryosu Raporu"])
    ws2.append(["Proje", project.name])
    ws2.append(["Repo", project.repo_url])
    ws2.append(["Üretildi", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")])
    ws2.append([""])
    ws2.append(["Toplam Senaryo", len(scenarios)])
    for stype in ("manual", "service", "automation"):
        count = sum(1 for s in scenarios if s.type == stype)
        ws2.append([stype.capitalize(), count])

    wb.save(out_path)


# ── Jira CSV (Xray / Zephyr Scale import formatı) ────────────────────────────

# Xray Test import CSV şeması (özelleştirilebilir)
_JIRA_HEADERS = [
    "Summary",          # Test başlığı
    "Issue Type",       # Test
    "Description",      # Gherkin veya adım açıklaması
    "Labels",           # tip:priority şeklinde etiket
    "Priority",         # Blocker / Critical / Major / Minor / Trivial
    "Component/s",      # Feature area
    "Test Type",        # Manual / Cucumber
    "Step Action",      # Manuel adım (tekrarlı satırlar)
    "Step Data",        # Test verisi
    "Step Result",      # Beklenen sonuç
]

_JIRA_PRIORITY_MAP = {
    "critical": "Blocker",
    "high": "Critical",
    "medium": "Major",
    "low": "Minor",
}

_JIRA_TEST_TYPE = {
    "manual": "Manual",
    "service": "Manual",
    "automation": "Cucumber",
}


def _to_jira_csv(
    project: NexusProject,
    scenarios: list[NexusScenario],
    cases_by_scenario: dict[str, list[NexusCase]],
) -> str:
    """Jira/Xray uyumlu CSV oluştur (UTF-8 BOM ile — Excel uyumu için)."""
    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_ALL)
    writer.writerow(_JIRA_HEADERS)

    for sc in scenarios:
        priority = _JIRA_PRIORITY_MAP.get(sc.priority, "Major")
        labels = f"nexus:{sc.type} priority:{sc.priority}"
        if sc.feature_area:
            labels += f" area:{sc.feature_area.replace(' ', '_')}"
        test_type = _JIRA_TEST_TYPE.get(sc.type, "Manual")

        sc_cases = cases_by_scenario.get(sc.id, [])

        if not sc_cases:
            # Case yok — tek satır
            description = sc.gherkin or sc.notes or ""
            writer.writerow([
                sc.title,
                "Test",
                description,
                labels,
                priority,
                sc.feature_area or "",
                test_type,
                "",  # Step Action
                "",  # Step Data
                "",  # Step Result
            ])
        else:
            first = True
            for case in sc_cases:
                steps = case.steps or []
                if not steps:
                    steps = [{"step": case.name, "expected": case.expected_result or "", "data": ""}]

                for step_idx, step in enumerate(steps):
                    if isinstance(step, dict):
                        action = step.get("step", "")
                        data = json.dumps(case.test_data, ensure_ascii=False) if (step_idx == 0 and case.test_data) else step.get("data", "")
                        result = step.get("expected", case.expected_result or "")
                    else:
                        action, data, result = str(step), "", ""

                    writer.writerow([
                        sc.title if first else "",
                        "Test" if first else "",
                        (sc.gherkin or sc.notes or "") if first else "",
                        labels if first else "",
                        priority if first else "",
                        (sc.feature_area or "") if first else "",
                        test_type if first else "",
                        action,
                        data,
                        result,
                    ])
                    first = False

    return "\ufeff" + output.getvalue()  # BOM ekle


# ── Ana İş Fonksiyonu ─────────────────────────────────────────────────────────

def run_export_job(export_id: str) -> None:
    """Arka planda çalıştırılacak export iş fonksiyonu."""
    db: Session = SessionLocal()
    try:
        export: Optional[NexusExport] = db.query(NexusExport).filter(NexusExport.id == export_id).first()
        if not export:
            _log.error("Export bulunamadı: %s", export_id)
            return

        export.status = "running"
        db.commit()

        project: Optional[NexusProject] = db.query(NexusProject).filter(NexusProject.id == export.project_id).first()
        if not project:
            export.status = "failed"
            db.commit()
            return

        # Senaryo filtresi
        q = db.query(NexusScenario).filter(NexusScenario.project_id == export.project_id)
        if export.scenario_ids:
            q = q.filter(NexusScenario.id.in_(export.scenario_ids))
        scenarios = q.all()

        if not scenarios:
            export.status = "failed"
            db.commit()
            _log.warning("Export %s: senaryo bulunamadı", export_id)
            return

        out_dir = _ensure_dir(_EXPORTS_BASE / export.project_id)
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        fmt = export.format

        if fmt == "gherkin":
            out_path = out_dir / f"scenarios_{timestamp}.feature"
            content = _to_gherkin(project, scenarios)
            out_path.write_text(content, encoding="utf-8")

        elif fmt == "postman":
            out_path = out_dir / f"collection_{timestamp}.json"
            collection = _to_postman(project, scenarios)
            out_path.write_text(json.dumps(collection, ensure_ascii=False, indent=2), encoding="utf-8")

        elif fmt in ("excel", "jira"):
            # Her iki format için case'leri yükle
            scenario_ids = [s.id for s in scenarios]
            all_cases = db.query(NexusCase).filter(NexusCase.scenario_id.in_(scenario_ids)).all()
            cases_by_scenario: dict[str, list[NexusCase]] = {}
            for c in all_cases:
                cases_by_scenario.setdefault(c.scenario_id, []).append(c)

            if fmt == "excel":
                out_path = out_dir / f"scenarios_{timestamp}.xlsx"
                _to_excel(project, scenarios, cases_by_scenario, out_path)
            else:
                out_path = out_dir / f"jira_import_{timestamp}.csv"
                csv_content = _to_jira_csv(project, scenarios, cases_by_scenario)
                out_path.write_text(csv_content, encoding="utf-8")

        else:
            export.status = "failed"
            db.commit()
            _log.error("Bilinmeyen export formatı: %s", fmt)
            return

        export.file_path = str(out_path)
        export.status = "done"
        db.commit()
        _log.info("Export %s tamamlandı: %s", export_id, out_path)

    except Exception as exc:
        _log.exception("Export %s başarısız: %s", export_id, exc)
        if export:
            export.status = "failed"
            try:
                db.commit()
            except Exception:
                pass
    finally:
        db.close()
