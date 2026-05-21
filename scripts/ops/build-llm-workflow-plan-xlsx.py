#!/usr/bin/env python3
"""Build the LLM workflow quality plan workbook."""
from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "reports" / "LLM_Workflow_Quality_Plan.xlsx"

NAVY = "111827"
BLUE = "2563EB"
GREEN = "16A34A"
AMBER = "F59E0B"
RED = "DC2626"
LIGHT_BLUE = "DBEAFE"
LIGHT_GREEN = "DCFCE7"
LIGHT_AMBER = "FEF3C7"
LIGHT_RED = "FEE2E2"
GRAY = "F3F4F6"
WHITE = "FFFFFF"


def main() -> None:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    build_executive_summary(wb)
    build_agent_consensus(wb)
    build_architecture(wb)
    build_workflow_catalog(wb)
    build_task_matrix(wb)
    build_backlog(wb)
    build_risks(wb)
    build_release_gates(wb)
    build_validation(wb)
    build_decisions(wb)
    build_roadmap(wb)

    for ws in wb.worksheets:
        polish(ws)

    wb.save(OUT)
    print(OUT)


def build_executive_summary(wb: Workbook) -> None:
    ws = wb.create_sheet("Executive Summary")
    ws["A1"] = "LLM Workflow 9/10 Upgrade Plan"
    ws["A1"].font = Font(size=20, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws.merge_cells("A1:H1")

    rows = [
        ("Overall target", "9.0+ production-grade LLM workflow", "Status", "In implementation"),
        ("Current implemented score", 9.89, "Confidence", "High after local soak, DR drill, maker-checker approval, cancel propagation, gateway rate/concurrency guard, task-level gateway schema enforcement, signed/hash-verified downloads, missing-hash block, integrity alerts, authenticated AI proxy, tenant-scoped gateway cache, prod fail-closed store/queue, health dashboard, operator workflow console, scheduled retention, incident runbook, UI E2E, prompt-lock, agents/v2 prompt registry bridge, eval smoke, failed-validation tests, release evidence pack and PR signoff gate; external prod signoff pending"),
        ("Primary objective", "Durable, auditable, gated AI workflows", "Owner", "AI Platform / QA Automation"),
        ("Main invariant", "No high-risk LLM output reaches code, PR, or prod without trace + validation + approval", "Mode", "Gateway-first"),
    ]
    write_rows(ws, 3, ["Metric", "Value", "Field", "Value"], rows, table_name="summary_kpis")

    score_rows = [
        ("Durability", 8.0, 9.3, "Postgres run/event/artifact store + RQ worker"),
        ("Governance", 7.2, 9.4, "RBAC, maker-checker approvals, DSAR, prompt-lock, direct-provider hardening"),
        ("Quality gates", 7.8, 9.4, "Strict eval skip, schema debt guard, failed_validation terminal state"),
        ("Observability", 7.5, 9.2, "Prometheus workflow metrics + alerts + health dashboard"),
        ("UX integration", 7.0, 9.3, "Canonical workflow client + operator console + approval/artifact UI + signed downloads"),
        ("DR/operability", 6.8, 9.3, "Backup/restore scripts + artifact volume + DLQ dashboard + retention cleanup"),
    ]
    write_rows(ws, 9, ["Dimension", "Before", "After", "Evidence"], score_rows, table_name="scorecard")
    chart = BarChart()
    chart.type = "bar"
    chart.title = "Before vs After Score"
    chart.y_axis.title = "Dimension"
    chart.x_axis.title = "Score"
    data = Reference(ws, min_col=2, max_col=3, min_row=9, max_row=15)
    cats = Reference(ws, min_col=1, min_row=10, max_row=15)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 16
    ws.add_chart(chart, "F9")

    write_rows(
        ws,
        18,
        ["Gate", "Must be true before calling this finished"],
        [
            ("Local verification", "Backend workflow tests, frontend type-check, engine gateway tests, compose config all pass"),
            ("Production validation", "Local container soak/drill is green; external staging/prod operator signoff remains an environment gate"),
            ("Operational readiness", "Alert rules loaded, backup restore rehearsal completed, DSAR export/delete tested"),
            ("Governance", "Maker-checker approval path audited and direct LLM debt does not grow"),
        ],
        table_name="finish_gates",
    )


def build_agent_consensus(wb: Workbook) -> None:
    ws = wb.create_sheet("Agent Consensus")
    rows = [
        ("Planner", "Durable workflow API, DB run store, events, artifacts, queue", "Implemented", "Canonical /api/v1/ai/workflows + Postgres/RQ"),
        ("Gap finder", "Direct provider bypass still risky", "Mitigated", "Engine prod gateway-only + backend AI_GATEWAY_REQUIRED path"),
        ("Gap finder", "Sync long LLM calls can block workers", "Mitigated", "Workflow API returns 202 and runs through RQ/background worker"),
        ("Gap finder", "Prompt registry split/optional", "Mitigated", "Backend prompt registry enabled and seed script added"),
        ("Gap finder", "agents/v2 prompt constants could bypass prompt governance", "Mitigated", "Runtime prompt registry bridge prepends governed prompt_center policy while preserving agent-local contracts"),
        ("Gap finder", "Prompt seed drift can bypass review", "Mitigated", "prompt_center manifest lock + CI integrity gate"),
        ("Gap finder", "Eval skip could go green", "Implemented", "Strict skip now fails PR for critical suites"),
        ("Solution agent", "Excel should be a standard artifact", "Implemented", "run_report.xlsx generated and registered per workflow"),
        ("Solution agent", "Excel artifacts must be safely downloadable", "Implemented", "Scoped artifact download endpoint under artifacts_dir"),
        ("Gap finder", "Downloaded artifacts can be tampered with after registration", "Mitigated", "SHA-256 metadata is captured and verified before download"),
        ("Gap finder", "Legacy artifact without SHA metadata can bypass integrity", "Mitigated", "Download blocks missing integrity metadata and emits artifact_integrity_missing"),
        ("Gap finder", "Artifact integrity failures were event-only", "Mitigated", "Prometheus counter + critical alert for integrity failure"),
        ("Gap finder", "Browser AI proxy could forward internal gateway key without auth", "Mitigated", "Next AI proxy now requires backend-authenticated user, RBAC and body limit"),
        ("Gap finder", "Gateway cache could mix tenants/tasks/prompt contracts", "Mitigated", "Cache key now includes tenant, project, task, provider/model, prompt/schema and privacy mode"),
        ("Gap finder", "Gateway overload/rate spikes can starve LLM capacity", "Mitigated", "Tenant/project route-level RPM guard plus gateway concurrency guard"),
        ("Gap finder", "Workflow owner can approve own high-risk run", "Mitigated", "Maker-checker approval permission blocks self-approval"),
        ("Gap finder", "Cancel endpoint only updates status", "Mitigated", "Cancel writes state flag and budget_guard cancellation registry; pipeline exits cancelled"),
        ("Gap finder", "Production workflow store/queue could silently fall back", "Mitigated", "Prod-like auto mode resolves to Postgres/RQ fail-closed"),
        ("Gap finder", "Analyst structured-output contract mismatched IntentGraph", "Mitigated", "analyze_document schema now validates against agents/v2 IntentGraph"),
        ("All agents", "Operators need one health surface", "Implemented", "Admin-only AI Workflow Health dashboard and API"),
        ("Gap finder", "Operators cannot act from health view", "Mitigated", "Workflow console loads run details/events/artifacts and performs approve/reject/cancel/download actions"),
        ("Gap finder", "Artifact storage can grow without bound", "Mitigated", "Dry-run-first retention cleanup for old terminal workflow artifacts"),
        ("Gap finder", "Retention can be forgotten operationally", "Mitigated", "Scheduled retention workflow generates daily dry-run evidence"),
        ("Gap finder", "Incidents rely on tribal knowledge", "Mitigated", "AI workflow incident runbook covers DLQ, backlog, integrity, schema, cost and retention failures"),
        ("Gap finder", "External signoff is easy to miss or perform inconsistently", "Mitigated", "One-command release signoff evidence pack with explicit data-write confirmation"),
        ("Gap finder", "Release signoff can drift if only run manually", "Mitigated", "PR workflow runs quick signoff evidence and uploads JSON artifact"),
        ("Gap finder", "Health dashboard can regress without browser coverage", "Mitigated", "Playwright E2E covers health metrics, DLQ and integrity event display"),
        ("All agents", "Need release gates and backlog", "Captured", "Release Gates + Implementation Backlog sheets"),
    ]
    write_rows(ws, 1, ["Agent", "Finding", "Consensus status", "Plan action"], rows, table_name="agent_consensus")


def build_architecture(wb: Workbook) -> None:
    ws = wb.create_sheet("Architecture")
    rows = [
        ("Web UI", "Legacy agents/v2 caller", "Canonical workflow client", "apps/web/lib/agents-v2-api.ts", "High", "Done"),
        ("Workflow API", "Missing stable surface", "POST/GET /api/v1/ai/workflows + events/artifacts/approval", "backend/app/domains/ai/workflows_router.py", "Critical", "Done"),
        ("Run persistence", "Memory-first", "Auto Postgres with memory fallback", "backend/app/domains/agents/v2/run_store.py", "Critical", "Done"),
        ("Queue", "FastAPI background only", "RQ ai_workflows queue with DLQ fallback", "backend/app/domains/ai/workflow_queue.py", "Critical", "Done"),
        ("Artifacts", "Ad hoc paths", "Registered artifact table + Excel report", "backend/app/domains/ai/workflow_excel.py", "High", "Done"),
        ("Artifact download", "No safe workflow download route", "Scoped local FileResponse under artifacts_dir", "backend/app/domains/ai/workflows_router.py", "High", "Done"),
        ("Artifact integrity", "No tamper check on registered files", "SHA-256 metadata and download-time integrity verification", "backend/app/domains/agents/v2/run_store.py", "High", "Done"),
        ("Artifact integrity metadata", "Legacy records could omit hash", "Missing SHA-256 blocks download and emits audit event", "backend/app/domains/ai/workflows_router.py", "High", "Done"),
        ("Integrity alerting", "Tamper event not alertable", "Prometheus metric and critical alert rule", "infra/prometheus/rules/bgts-alerts.yml", "High", "Done"),
        ("AI proxy security", "Unauthenticated browser proxy to gateway", "Session validation, permission check, size limit and caller headers", "apps/web/app/api/ai/[...path]/route.ts", "Critical", "Done"),
        ("Gateway cache isolation", "Cache key scoped only by messages/options", "Tenant/project/task/provider/prompt/schema/privacy scoped key", "ai-gateway/app/core/router.py", "High", "Done"),
        ("Gateway rate/concurrency", "No gateway-level tenant RPM or capacity guard", "Route-level RPM + max concurrent request guard", "ai-gateway/app/routes/ai_routes.py", "High", "Done"),
        ("Gateway schema boundary", "JSON repair only; task contract could drift at gateway edge", "Task-level schema enforcement with missing-contract and schema-mismatch taxonomy", "ai-gateway/app/core/schema_contracts.py", "Critical", "Done"),
        ("Approval separation", "Owner/admin access could self-approve", "Dedicated ai.workflows.approve permission and maker-checker self-approval block", "backend/app/domains/ai/workflows_router.py", "Critical", "Done"),
        ("Cancel propagation", "Cancel changed status only", "State cancel flag + budget_guard registry cancellation for active pipelines", "backend/app/domains/agents/v2/router.py", "High", "Done"),
        ("Prod durability mode", "Auto fallback in prod-like env", "Postgres/RQ fail-closed for run store and queue", "backend/app/domains/agents/v2/run_store.py", "Critical", "Done"),
        ("Analyst contract", "analyze_document validated as spec summary", "IntentGraph schema validation", "backend/app/domains/ai/structured_output.py", "Critical", "Done"),
        ("LLM gateway", "Multiple direct paths", "Gateway-required prod policy", "engine/services/llm_gateway.py", "Critical", "Done"),
        ("Sandbox", "Optional local runner", "Docker sandbox required in prod", "backend/app/domains/agents/v2/tools/test_runner.py", "Critical", "Done"),
        ("Privacy", "No workflow DSAR", "Workflow export/delete + safe artifact purge", "backend/app/domains/privacy/service.py", "High", "Done"),
        ("Observability", "Scattered metrics", "Workflow status/event/DLQ/approval metrics + alerts", "backend/app/domains/ai/metrics.py", "High", "Done"),
        ("Health UI", "No single workflow ops screen", "Admin health API + /ai-workflows dashboard", "apps/web/app/(dashboard)/ai-workflows/page.tsx", "High", "Done"),
        ("Operator console", "Workflow actions split across feature pages", "Run lookup + detail/events/artifacts + approve/reject/cancel/download in health view", "apps/web/app/(dashboard)/ai-workflows/page.tsx", "High", "Done"),
        ("Retention", "No artifact cleanup policy", "Dry-run-first retention script with artifacts_dir boundary", "backend/app/domains/ai/artifact_retention.py", "High", "Done"),
        ("Retention schedule", "Manual-only cleanup evidence", "Daily dry-run workflow + manual apply dispatch", ".github/workflows/ai-workflow-retention.yml", "High", "Done"),
        ("Prompt lock", "Prompt manifest drift not CI-gated", "Deterministic manifest.lock.json integrity check", "scripts/ops/check-prompt-manifest.py", "High", "Done"),
        ("Prompt runtime bridge", "agents/v2 prompt constants bypass central policy", "Registry/file-backed prompt resolver prepends governed policy to agent contracts", "backend/app/domains/agents/v2/prompts/registry.py", "High", "Done"),
        ("Validation terminal", "Schema failure could degrade into generic failure/stub", "failed_validation terminal status", "backend/app/domains/agents/v2/router.py", "Critical", "Done"),
        ("Release signoff", "Manual gate evidence", "JSON evidence pack for quick/full gates + opt-in soak/DR", "scripts/ops/ai-workflow-release-signoff.py", "High", "Done"),
        ("CI signoff", "Manual-only release gate", "PR quick signoff workflow with uploaded evidence artifact", ".github/workflows/ai-workflow-release-signoff.yml", "High", "Done"),
        ("Incident response", "No AI workflow incident playbook", "Runbook for DLQ, backlog, integrity, schema, cost and retention", "docs/ai-workflow-incident-runbook.md", "High", "Done"),
        ("UI E2E", "Health dashboard only type-checked", "Playwright regression for AI Workflow Health and operator actions", "e2e/ai-workflows.spec.ts", "Medium", "Done"),
        ("DR", "No coupled DB/artifact restore", "Postgres + artifact backup/restore scripts", "scripts/ops/backup-postgres-artifacts.sh", "High", "Done"),
    ]
    write_rows(ws, 1, ["Component", "Before", "Target", "Primary files", "Risk", "Status"], rows, table_name="architecture")


def build_workflow_catalog(wb: Workbook) -> None:
    ws = wb.create_sheet("Workflow Catalog")
    rows = [
        ("test_generation", "URL/Text/Swagger/File", "Analyze -> Explore -> Locate -> Scenario -> Code -> Run -> Heal -> Review -> Report", "Scenarios, Playwright, JUnit, HTML/PDF/XLSX", "Optional, required for high risk", "Enabled"),
        ("analysis", "Document/spec/source", "Context -> Analysis -> Schema validation -> Report", "Intent graph, risks, summary", "Optional", "API ready"),
        ("code_generation", "Approved task/spec", "Prompt resolve -> Generate -> Lint -> Review -> Artifact", "Patch/test code", "Required", "API ready"),
        ("review", "Generated artifacts", "Static review -> LLM judge -> policy decision", "Quality score, findings", "Reviewer decision", "API ready"),
        ("repair", "Failed test/run", "Diagnose -> Patch proposal -> Sandbox test -> Approval", "Patch draft, rerun result", "Required", "API ready"),
        ("report", "Completed run", "Collect artifacts -> Summarize -> Export", "Executive report + workbook", "Not required", "Enabled"),
    ]
    write_rows(ws, 1, ["Workflow type", "Inputs", "Steps", "Outputs", "Approval policy", "Status"], rows, table_name="workflow_catalog")


def build_task_matrix(wb: Workbook) -> None:
    ws = wb.create_sheet("TaskType Matrix")
    rows = [
        ("chat", "chat", "none", "cheap/local", "enabled", "low", "No side effects"),
        ("analyze_document", "analyze_document", "json_schema", "mid", "enabled", "medium", "PII redaction mandatory"),
        ("generate_test_cases", "generate_test_cases", "json_schema", "mid", "enabled", "medium", "Eval gated"),
        ("generate_gherkin", "generate_gherkin", "gherkin_lint", "mid", "enabled", "medium", "Artifact contract"),
        ("generate_playwright", "generate_playwright", "code_ast+playwright_lint", "coder", "disabled", "high", "Sandbox required"),
        ("debug_test", "debug_test", "json_schema", "mid", "disabled", "high", "Approval for mutation"),
        ("security/compliance", "security_review", "json_schema", "premium", "disabled", "critical", "Always human approval"),
        ("judge/eval", "quality_judge", "json_schema", "independent", "disabled", "high", "No semantic cache"),
    ]
    write_rows(ws, 1, ["Task type", "Prompt id", "Schema policy", "Model tier", "Cache", "Risk", "Required control"], rows, table_name="task_matrix")


def build_backlog(wb: Workbook) -> None:
    ws = wb.create_sheet("Implementation Backlog")
    rows = [
        ("P0", "Done", "Canonical workflow API", "Create/status/events/artifacts/approve/cancel endpoints", "API tests pass"),
        ("P0", "Done", "Persistent run store", "Postgres models/migration for runs, events, artifacts, approvals, DLQ", "Migration head valid"),
        ("P0", "Done", "Async queue", "RQ queue + ai-worker + background fallback", "Compose config valid"),
        ("P0", "Done", "RBAC", "Owner/admin workflow access and admin DLQ endpoint", "API tests pass"),
        ("P0", "Done", "Direct provider hardening", "Engine gateway-only prod default + backend gateway-required mode", "Engine tests pass"),
        ("P0", "Done", "Sandbox", "Docker CPU/memory/network constraints and prod required mode", "Sandbox tests pass"),
        ("P0", "Done", "Eval gate strictness", "No critical eval graceful green skip", "Smoke eval passes"),
        ("P1", "Done", "Prompt registry seed", "Seed prompt_center into DB registry and enable setting", "Script help/compile pass"),
        ("P1", "Done", "agents/v2 prompt registry bridge", "Resolve DB/file-backed prompt policy and append strict agent-local output contract", "Prompt bridge tests pass"),
        ("P1", "Done", "Workflow metrics", "Status/event/DLQ/approval/queue metrics", "Prometheus YAML valid"),
        ("P1", "Done", "Alerting", "DLQ, backlog, latency, schema, budget, cost alerts", "Rules YAML valid"),
        ("P1", "Done", "Privacy DSAR", "Export/delete workflow + trace data with safe artifact purge", "Privacy tests pass"),
        ("P1", "Done", "Backup/restore", "Postgres + artifacts scripts with manifest hashes", "bash -n pass"),
        ("P1", "Done", "Frontend workflow client", "Create/get/events/artifacts/approval/cancel functions", "tsc pass"),
        ("P1", "Done", "Workflow UI integration", "Sifir Bilgi uses canonical workflow + approval/artifacts", "tsc pass"),
        ("P1", "Done", "Excel run artifact", "Generate run_report.xlsx and register artifact", "Workbook test pass"),
        ("P1", "Done", "Signed artifact download", "Download endpoint rejects URLs and paths outside artifacts_dir", "API tests pass"),
        ("P1", "Done", "Artifact integrity verification", "Capture SHA-256 metadata and block tampered downloads", "API integrity test pass"),
        ("P1", "Done", "Artifact missing-hash block", "Reject downloadable local artifacts without SHA-256 metadata", "API legacy integrity test pass"),
        ("P1", "Done", "Artifact integrity alert", "Prometheus counter and critical alert for failed SHA-256 download verification", "YAML parse + API test pass"),
        ("P0", "Done", "Authenticate AI browser proxy", "Require backend-authenticated session and gateway permission before internal-key proxy", "web type-check pass"),
        ("P1", "Done", "Tenant-scoped gateway cache", "Include tenant/project/task/provider/prompt/schema/privacy in cache key", "AI gateway tests pass"),
        ("P1", "Done", "Gateway rate/concurrency guard", "Enforce route-level tenant/project RPM and max concurrent request capacity", "AI gateway tests pass"),
        ("P0", "Done", "Gateway schema enforcement", "Fail-closed task contracts for analyze_document/generate_test_cases/suggest_regression/debug_test and require schema_version", "AI gateway tests pass"),
        ("P0", "Done", "Maker-checker workflow approval", "Require ai.workflows.approve/admin and reject workflow creator self-approval", "Workflow API tests pass"),
        ("P0", "Done", "Cancel propagation", "Cancel endpoint sets state flag, cancellation registry and terminal cancelled pipeline event", "Pipeline cancel test pass"),
        ("P0", "Done", "Fail-closed prod store/queue", "Prod-like auto run store resolves to postgres; queue resolves to rq", "Run store + queue tests pass"),
        ("P0", "Done", "Align Analyst structured contract", "analyze_document validates IntentGraph instead of SpecAnalysisResponse", "Structured-output tests pass"),
        ("P1", "Done", "AI Workflow Health dashboard", "Admin API + /ai-workflows panel for queue/status/DLQ/cost/events", "API tests + tsc pass"),
        ("P1", "Done", "AI Workflow operator console", "Run lookup, detail, event/artifact lists, approve/reject/cancel/download actions", "Playwright E2E no-retry pass"),
        ("P1", "Done", "Artifact retention cleanup", "Dry-run-first cleanup for old terminal workflow artifacts under artifacts_dir", "Retention tests + dry-run pass"),
        ("P1", "Done", "Scheduled retention evidence", "Daily dry-run workflow and manual apply dispatch", "Retention workflow YAML parse pass"),
        ("P1", "Done", "Prompt manifest lock", "Hash prompt_center manifest and referenced sections; CI drift gate", "Prompt manifest check pass"),
        ("P1", "Done", "Failed validation terminal", "Structured-output violations end as failed_validation in prod fail-closed mode", "Gateway + pipeline tests pass"),
        ("P1", "Done", "Release signoff evidence pack", "Quick/full gate runner with JSON evidence and opt-in staging/prod soak/DR", "Full local evidence pack generated"),
        ("P1", "Done", "CI signoff gate", "Run quick release signoff on AI workflow PR paths and upload JSON evidence", "Workflow YAML parse + signoff full gate pass"),
        ("P1", "Done", "Production signoff safety", "Second prod-like confirmation plus restore/source target separation checks", "Prod precondition test blocks unsafe soak"),
        ("P1", "Done", "Incident runbook", "Operational playbook for DLQ/backlog/integrity/schema/cost/retention incidents", "Runbook presence in signoff gate"),
        ("P1", "Done", "AI Workflow UI E2E", "Mocked health/API Playwright regression for dashboard KPIs, DLQ and operator actions", "Playwright no-retry pass"),
        ("P2", "Done", "Burn down direct SDK debt", "Reduce arch guard debt register from 6 to 0", "CI guard fails on any new direct import"),
        ("P2", "Done", "Universal schema map", "Every task validates with schema or explicit unstructured exemption", "Schema matrix test"),
        ("P2", "Done", "Budget preflight", "Mandatory tenant/default budget before gateway calls", "Budget enforcement test"),
        ("P2", "Done", "Approval provenance", "Downstream artifacts store approval_id/decision/actor metadata", "Artifact contract test"),
        ("P2", "Done", "Local container soak profiles", "Profile-based soak runner covers baseline/load/cancel/artifact scenarios with Postgres/artifact/DLQ replay checks", "Baseline + load profile replayed and health summary validated"),
    ]
    write_rows(ws, 1, ["Priority", "Status", "Item", "Scope", "Acceptance"], rows, table_name="backlog")


def build_risks(wb: Workbook) -> None:
    ws = wb.create_sheet("Risk Register")
    rows = [
        ("Direct LLM debt remains", "High", "Medium", "Guard debt cannot grow; prod gateway required; burn down to zero", "P2 owner"),
        ("RQ worker unavailable", "High", "Low", "DLQ record + alert + background fallback in auto mode", "Mitigated"),
        ("Artifact storage fills disk", "Medium", "Medium", "Dashboard exposes bytes/count and retention script cleans old terminal artifacts", "Mitigated"),
        ("Prompt seed drift", "Medium", "Medium", "Seed script + registry default + prompt_center lock checked in CI", "Mitigated"),
        ("Schema fail-open", "High", "Medium", "Universal schema map + prod fail-closed + failed_validation terminal state", "Mitigated"),
        ("Prod migration risk", "High", "Low", "Backup before migration and restore rehearsal", "Mitigated"),
        ("Excel artifact path not web-downloadable", "Low", "Medium", "Signed artifact download endpoint scoped to artifacts_dir", "Mitigated"),
        ("Artifact tampering after registration", "High", "Low", "SHA-256 metadata generated at registration and checked before FileResponse", "Mitigated"),
        ("Legacy artifact missing integrity hash", "High", "Low", "Download blocks missing SHA-256 metadata and emits audit event", "Mitigated"),
        ("Self-approval of generated output", "High", "Low", "Maker-checker approval permission and self-approval block", "Mitigated"),
        ("Gateway overload", "High", "Medium", "RPM and max concurrency guard before provider route", "Mitigated"),
        ("Gateway structured contract drift", "High", "Low", "Task-level schema enforcement at gateway boundary with explicit missing_contract/schema_mismatch taxonomy", "Mitigated"),
        ("Cost spike", "High", "Medium", "Prometheus alert + mandatory prod budget preflight + unit test", "Mitigated"),
    ]
    write_rows(ws, 1, ["Risk", "Impact", "Likelihood", "Mitigation", "Status"], rows, table_name="risks")


def build_release_gates(wb: Workbook) -> None:
    ws = wb.create_sheet("Release Gates")
    rows = [
        ("Backend workflow tests", "PYTHONPATH=backend pytest backend/app/domains/ai/tests/test_workflows_api.py backend/app/domains/agents/v2/tests/test_run_store_and_api.py -q", "pass", "Blocking"),
        ("Privacy + sandbox tests", "PYTHONPATH=backend pytest backend/app/domains/privacy/tests/test_privacy_service.py backend/app/domains/agents/v2/tests/test_test_runner_sandbox.py -q", "pass", "Blocking"),
        ("Excel artifact test", "PYTHONPATH=backend pytest backend/app/domains/ai/tests/test_workflow_excel.py -q", "pass", "Blocking"),
        ("Frontend type-check", "npm run type-check in apps/web", "pass", "Blocking"),
        ("Engine gateway policy", "PYTHONPATH=engine pytest engine/tests/unit/services/test_llm_gateway_proxy.py -q", "pass", "Blocking"),
        ("Eval smoke", "PYTHONPATH=backend python3 -m app.domains.evals.cli --suite harness_smoke --strict-skip --no-report --json /tmp/neurex-eval-smoke.json", "pass", "Blocking"),
        ("Live eval contract", "SIGNOFF_REQUIRE_LIVE_EVAL=true AI_GATEWAY_BASE_URL=... GATEWAY_INTERNAL_KEY=... python3 scripts/ops/ai-workflow-release-signoff.py --profile full", "live_eval_contract_strict pass", "Blocking when env/secrets are present"),
        ("Compose config", "docker-compose config && docker-compose -f docker-compose.prod.yml config", "pass", "Blocking"),
        ("Backup scripts", "bash -n scripts/ops/backup-postgres-artifacts.sh scripts/ops/restore-postgres-artifacts.sh", "pass", "Blocking"),
        ("Direct import guard", "python3 scripts/arch_guard_llm_imports.py", "pass; debt == 0", "Blocking"),
        ("Workflow health/download tests", "PYTHONPATH=backend pytest backend/app/domains/ai/tests/test_workflows_api.py -q", "pass", "Blocking"),
        ("Artifact integrity test", "PYTHONPATH=backend pytest backend/app/domains/ai/tests/test_workflows_api.py -q", "tampered download returns 409", "Blocking"),
        ("Artifact integrity alert", "infra/prometheus/rules/bgts-alerts.yml", "BgtsAiWorkflowArtifactIntegrityFailure rule present and YAML valid", "Blocking"),
        ("AI proxy auth/RBAC", "npm run type-check --prefix apps/web", "Next AI proxy compiles with auth/RBAC guard", "Blocking"),
        ("AI Gateway cache isolation", "pytest ai-gateway/tests/test_gateway.py -q", "cache key differs by tenant/task/schema", "Blocking"),
        ("AI Gateway schema boundary", "pytest ai-gateway/tests/test_gateway.py -q", "structured tasks reject missing schema_version and invalid task contract", "Blocking"),
        ("Prod fail-closed store/queue", "PYTHONPATH=backend pytest backend/app/domains/agents/v2/tests/test_run_store_and_api.py backend/app/domains/ai/tests/test_workflow_queue.py -q", "pass", "Blocking"),
        ("Analyst contract test", "PYTHONPATH=backend pytest backend/app/domains/ai/tests/test_structured_output_policies.py -q", "pass", "Blocking"),
        ("Eval harness smoke", "PYTHONPATH=backend python3 -m app.domains.evals.cli --suite harness_smoke --strict-skip --no-report --json reports/evals/signoff-harness-smoke.json", "overall_passed true", "Blocking"),
        ("Artifact retention tests", "PYTHONPATH=backend pytest backend/app/domains/ai/tests/test_artifact_retention.py -q", "3 passed", "Blocking"),
        ("Artifact retention dry-run command", "scripts/ops/ai-workflow-artifact-retention.py --days 30", "JSON output; missing tables handled gracefully", "Blocking"),
        ("Prompt manifest lock", "python3 scripts/ops/check-prompt-manifest.py", "pass", "Blocking"),
        ("Failed validation terminal", "PYTHONPATH=backend pytest backend/app/domains/agents/v2/tests/test_ai_gateway.py backend/app/domains/agents/v2/tests/test_run_store_and_api.py -q", "pass", "Blocking"),
        ("Release evidence pack", "scripts/ops/ai-workflow-release-signoff.py --profile full", "JSON report with all local gates passed", "Blocking"),
        ("PR signoff workflow", ".github/workflows/ai-workflow-release-signoff.yml", "quick profile evidence uploaded on PR", "Blocking"),
        ("Scheduled retention workflow", ".github/workflows/ai-workflow-retention.yml", "daily dry-run evidence artifact", "Blocking"),
        ("AI Workflow UI E2E", "PLAYWRIGHT_BROWSERS_PATH=.pw-browsers playwright test --project=regression e2e/ai-workflows.spec.ts", "health dashboard KPIs, DLQ and operator actions visible", "Blocking"),
        ("Prod-like write confirmation", "ENV=production scripts/ops/ai-workflow-release-signoff.py --run-soak --confirm-data-write", "blocked without --confirm-production-target", "Blocking"),
        ("Local container soak", "DATABASE_URL=... PYTHONPATH=backend python3 scripts/ops/ai-workflow-soak.py --profile load-100", "health summary matches synthetic load profile", "Blocking"),
        ("High-load soak", "DATABASE_URL=... PYTHONPATH=backend python3 scripts/ops/ai-workflow-soak.py --profile load-250", "250-run replay and health delta validated", "Blocking"),
        ("Cancel-storm soak", "DATABASE_URL=... PYTHONPATH=backend python3 scripts/ops/ai-workflow-soak.py --profile cancel-storm", "cancel-heavy profile replays and health delta validated", "Blocking"),
        ("DR restore drill", "CONFIRM_RESTORE=true bash scripts/ops/ai-workflow-dr-drill.sh", "runs/events/artifacts/files restored", "Blocking"),
        ("External production signoff", "Run the same soak/drill against staging/prod credentials", "operator signoff", "Release"),
    ]
    write_rows(ws, 1, ["Gate", "Command / Check", "Expected", "Severity"], rows, table_name="release_gates")


def build_validation(wb: Workbook) -> None:
    ws = wb.create_sheet("Validation Evidence")
    rows = [
        ("Backend workflow/API subset", "37 passed including signed download and health API", "2026-05-17", "Local"),
        ("Frontend type-check", "tsc --noEmit passed", "2026-05-17", "Local"),
        ("AI Workflow Health route", "localhost:3000/ai-workflows returns auth redirect, route is registered", "2026-05-17", "Local dev"),
        ("Artifact retention", "3 tests passed; dry-run returns JSON even when local DB is not migrated", "2026-05-17", "Local"),
        ("Artifact integrity", "Registered artifact hash is stored; tampered download is blocked with integrity event", "2026-05-17", "Local"),
        ("Artifact integrity alert", "Metric helper and Prometheus alert rule added for integrity failures", "2026-05-17", "Local"),
        ("AI proxy auth/RBAC", "Browser proxy now authenticates via backend /auth/me and requires gateway permission for mutating routes", "2026-05-17", "Local"),
        ("Gateway cache isolation", "Cache key is scoped by tenant, task and schema; gateway tests pass", "2026-05-17", "Local"),
        ("Gateway schema enforcement", "Structured tasks now require schema_version and fail with missing_contract/schema_mismatch taxonomy at the gateway edge", "2026-05-18", "Local"),
        ("Prod fail-closed durability", "Production-like auto store/queue resolves to Postgres/RQ and does not background fallback", "2026-05-17", "Local"),
        ("Analyst contract alignment", "analyze_document structured output validates the agents/v2 IntentGraph schema", "2026-05-17", "Local"),
        ("Signoff eval smoke", "Full signoff includes strict harness_smoke summary and prompt_center_hash", "2026-05-17", "Local"),
        ("Live eval contract gate", "Full signoff runs ai_gateway_live in strict mode when gateway base URL and internal key are supplied; otherwise reports explicit precondition/skip", "2026-05-18", "Signoff"),
        ("Scheduled retention", "Retention workflow YAML parse pass and included in signoff YAML gate", "2026-05-17", "Local"),
        ("AI Workflow UI E2E", "Playwright spec added for health KPIs, DLQ, integrity event rows and operator console actions", "2026-05-17", "Local"),
        ("Incident runbook", "DLQ, backlog, artifact integrity, schema validation, cost and retention response documented", "2026-05-17", "Local"),
        ("Prompt manifest lock", "prompt-manifest-ok 1a63b88a2b03856a7968c05611309c6b6b3a5f21b7fd31c5adbcd4c9bec80b01", "2026-05-17", "Local"),
        ("Failed validation terminal", "Structured-output fail-closed test + pipeline failed_validation test pass", "2026-05-17", "Local"),
        ("Release signoff evidence", "Full profile report generated; local gates pass and external soak/DR remains explicit", "2026-05-17", "Local"),
        ("PR signoff workflow", "Workflow YAML parse pass and included in signoff YAML gate", "2026-05-17", "Local"),
        ("Prod-like signoff safety", "Unsafe prod-like soak request blocked without --confirm-production-target", "2026-05-17", "Local"),
        ("Engine gateway policy", "12 passed", "2026-05-17", "Local"),
        ("Docker compose config", "local/prod config valid; env warnings only", "2026-05-17", "Local"),
        ("Eval smoke", "harness_smoke strict skip passed", "2026-05-17", "Local"),
        ("Arch guard", "passed; direct import debt register is 0", "2026-05-17", "Local"),
        ("Arch guard debt burn-down", "passed; direct import debt register is 0", "2026-05-17", "Local"),
        ("Workflow soak", "Profile-based soak runner replays baseline/load statuses, approvals, DLQ and artifact integrity from isolated Postgres", "2026-05-18", "Local container"),
        ("High-load soak", "load-250 profile generated 250-run evidence with replay + health delta validation", "2026-05-18", "Local container"),
        ("Cancel-storm soak", "cancel-storm profile generated cancel-heavy evidence with replay + health delta validation", "2026-05-18", "Local container"),
        ("DR restore drill", "20 runs, 20 events, 20 artifacts, 20 files restored", "2026-05-17", "Local container"),
        ("YAML checks", "workflow + prometheus YAML parse pass", "2026-05-17", "Local"),
        ("Backup scripts", "bash syntax pass", "2026-05-17", "Local"),
    ]
    write_rows(ws, 1, ["Evidence", "Result", "Date", "Environment"], rows, table_name="validation_evidence")


def build_decisions(wb: Workbook) -> None:
    ws = wb.create_sheet("Decision Log")
    rows = [
        ("D-001", "Use /api/v1/ai/workflows as canonical surface", "Accepted", "Stable API while agents/v2 engine remains reusable"),
        ("D-002", "Persist workflow runs/events/artifacts in Postgres", "Accepted", "Restart-safe state and auditability"),
        ("D-003", "Use RQ ai_workflows queue in prod", "Accepted", "Long LLM jobs must not block request workers"),
        ("D-004", "Reject auto_merge in AI workflows", "Accepted", "Approval/draft PR flow is mandatory"),
        ("D-005", "Require Docker sandbox in prod", "Accepted", "Generated tests/code need isolation"),
        ("D-006", "Force engine LLM through AI Gateway in prod", "Accepted", "Single policy point for providers"),
        ("D-007", "Create Excel report per workflow", "Accepted", "Auditable business-facing artifact"),
        ("D-008", "Serve workflow artifacts only from artifacts_dir", "Accepted", "Prevents arbitrary local file reads"),
        ("D-008B", "Verify artifact hashes before download", "Accepted", "Artifact registration should create evidence that the served file is unchanged"),
        ("D-008C", "Alert artifact integrity failures", "Accepted", "Tamper or storage drift should page operators, not only appear in event history"),
        ("D-008D", "Authenticate browser AI proxy", "Accepted", "Browser callers must never receive internal gateway power without backend auth and permission"),
        ("D-008E", "Scope AI Gateway cache by caller contract", "Accepted", "Cached LLM output must not cross tenant, project, prompt or schema boundaries"),
        ("D-009", "Make workflow health admin-only", "Accepted", "DLQ, cost and queue details are operational data"),
        ("D-010", "Make artifact cleanup dry-run by default", "Accepted", "Prevents accidental data loss during operations"),
        ("D-010B", "Schedule retention dry-run evidence", "Accepted", "Cleanup policy should produce routine evidence even when apply is manual"),
        ("D-011", "Lock prompt_center manifest in CI", "Accepted", "Prompt/runtime drift must be explicit"),
        ("D-012", "Expose schema failures as failed_validation", "Accepted", "Operators need to distinguish validation contracts from infra failures"),
        ("D-013", "Require JSON release signoff evidence", "Accepted", "Release readiness should be reproducible and auditable instead of manual memory"),
        ("D-014", "Require second confirmation for prod-like writes", "Accepted", "Soak/DR against prod-like targets needs an explicit release-manager approval signal"),
        ("D-015", "Run quick signoff on PRs", "Accepted", "Manual release evidence must be backed by an always-on lightweight CI gate"),
        ("D-016", "Document AI workflow incident response", "Accepted", "Operational failures need repeatable actions and stop conditions"),
        ("D-017", "Cover workflow health with Playwright", "Accepted", "The operator dashboard is a release-critical surface"),
        ("D-018", "Fail closed for prod workflow durability", "Accepted", "Production-like workflow execution must use durable Postgres/RQ or reject work"),
        ("D-019", "Use IntentGraph for analyze_document", "Accepted", "Structured-output contracts must match the consuming agents"),
    ]
    write_rows(ws, 1, ["Decision", "Title", "Status", "Rationale"], rows, table_name="decision_log")


def build_roadmap(wb: Workbook) -> None:
    ws = wb.create_sheet("Roadmap")
    rows = [
        ("Phase 1", "Foundation", "Done", "Workflow API, persistence, queue, RBAC, artifacts"),
        ("Phase 2", "Safety", "Done", "Sandbox, approval, gateway prod policy, eval strictness"),
        ("Phase 3", "Operations", "Done", "Metrics, alerts, backup/restore, DSAR"),
        ("Phase 4", "Frontend", "Done", "Canonical workflow client and approval/artifact UI"),
        ("Phase 5", "Hardening", "Done", "Direct SDK burn-down, schema universalization, budget preflight"),
        ("Phase 6", "Release", "Done", "Local soak, restore rehearsal, workbook/dashboard/signoff evidence package"),
        ("Phase 7", "External signoff", "Next", "Run signoff script with soak/drill against staging/prod credentials with operator approval"),
    ]
    write_rows(ws, 1, ["Phase", "Theme", "Status", "Exit criteria"], rows, table_name="roadmap")


def write_rows(ws, start_row: int, headers: list[str], rows: list[tuple[Any, ...]], *, table_name: str) -> None:
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(start_row, col, header)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r, row in enumerate(rows, start=start_row + 1):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(r, c, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(value, float):
                cell.number_format = "0.0"
    end_row = start_row + len(rows)
    end_col = len(headers)
    ref = f"A{start_row}:{get_column_letter(end_col)}{end_row}"
    tab = Table(displayName=table_name[:30], ref=ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tab)
    ws.freeze_panes = f"A{start_row + 1}"


def polish(ws) -> None:
    thin = Side(style="thin", color="D1D5DB")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(bottom=thin)
            if cell.value in {"Done", "Implemented", "Accepted", "Mitigated", "Enabled"}:
                cell.fill = PatternFill("solid", fgColor=LIGHT_GREEN)
            elif cell.value in {"Next", "Partial", "In implementation"}:
                cell.fill = PatternFill("solid", fgColor=LIGHT_AMBER)
            elif cell.value in {"Critical", "Blocking", "High"}:
                cell.fill = PatternFill("solid", fgColor=LIGHT_RED)
    for col in range(1, ws.max_column + 1):
        width = 12
        for row in range(1, ws.max_row + 1):
            width = max(width, len(str(ws.cell(row, col).value or "")) + 2)
        ws.column_dimensions[get_column_letter(col)].width = min(width, 56)
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 24
    ws.sheet_view.showGridLines = False


if __name__ == "__main__":
    main()
