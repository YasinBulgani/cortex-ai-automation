"""
core/excel_reporter.py — Excel Test Raporu Ureticisi

NexusQA projesindeki ExcelReportGenerator.java pattern'inin
openpyxl ile Python implementasyonu.

Ozellikler:
  - Test sonuclarini .xlsx formatinda raporlar
  - Basari/basarisizlik istatistikleri
  - Domain bazli sayfa ayirimi (multi-domain destegi)
  - Hucre renklendirme (yesil=passed, kirmizi=failed, sari=skipped)

Kullanim:
    reporter = ExcelReporter("Sprint-1 Raporu")
    reporter.add_result("Login Testi", "passed", 1200, "")
    reporter.add_result("Odeme Testi", "failed", 3500, "Timeout hatasi")
    path = reporter.save()
"""
from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False
    logger.warning("openpyxl yuklu degil; Excel rapor olusturulamayacak. pip install openpyxl")

from config.settings import settings


_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") if _HAS_OPENPYXL else None
_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") if _HAS_OPENPYXL else None
_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") if _HAS_OPENPYXL else None
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid") if _HAS_OPENPYXL else None
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11) if _HAS_OPENPYXL else None
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
) if _HAS_OPENPYXL else None

_STATUS_FILLS = {
    "passed": _GREEN,
    "failed": _RED,
    "skipped": _YELLOW,
    "error": _RED,
    "pending": _YELLOW,
}

_HEADERS = ["#", "Test Adi", "Durum", "Sure (ms)", "Domain", "Hata Mesaji", "Tarih"]


class ExcelReporter:
    """Test sonuclarini Excel dosyasina raporlar."""

    def __init__(self, report_name: str = "test_report", domain: str = ""):
        if not _HAS_OPENPYXL:
            raise ImportError("openpyxl paketi gerekli: pip install openpyxl")
        self.report_name = report_name
        self.domain = domain
        self._results: list[dict] = []
        self._start_time = datetime.now()

    def add_result(
        self,
        test_name: str,
        status: str,
        duration_ms: int = 0,
        error_message: str = "",
        domain: str = "",
    ):
        """Bir test sonucu ekler."""
        self._results.append({
            "test_name": test_name,
            "status": status.lower(),
            "duration_ms": duration_ms,
            "error_message": error_message,
            "domain": domain or self.domain,
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        })

    def save(self, output_dir: Path | str | None = None) -> str:
        """Excel dosyasini kaydeder ve dosya yolunu doner."""
        out = Path(output_dir) if output_dir else settings.REPORTS_DIR
        out.mkdir(parents=True, exist_ok=True)

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{self.report_name}_{ts}.xlsx"
        filepath = out / filename

        wb = Workbook()
        ws = wb.active
        ws.title = "Test Sonuclari"

        self._write_summary(ws)
        self._write_header(ws, start_row=4)
        self._write_results(ws, start_row=5)
        self._auto_width(ws)

        # Domain bazli ek sayfalar (birden fazla domain varsa)
        domains = {r["domain"] for r in self._results if r["domain"]}
        if len(domains) > 1:
            for domain in sorted(domains):
                dws = wb.create_sheet(title=domain[:31])
                domain_results = [r for r in self._results if r["domain"] == domain]
                self._write_header(dws, start_row=1)
                self._write_results(dws, start_row=2, results=domain_results)
                self._auto_width(dws)

        wb.save(str(filepath))
        logger.info("Excel rapor olusturuldu: %s", filepath)
        return str(filepath)

    # -- Internal helpers -------------------------------------------------------

    def _write_summary(self, ws):
        total = len(self._results)
        passed = sum(1 for r in self._results if r["status"] == "passed")
        failed = sum(1 for r in self._results if r["status"] == "failed")
        skipped = total - passed - failed
        rate = round(passed / total * 100, 1) if total else 0

        ws["A1"] = f"Rapor: {self.report_name}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"Tarih: {self._start_time.strftime('%d.%m.%Y %H:%M')}"
        ws["C2"] = f"Toplam: {total}"
        ws["D2"] = f"Basarili: {passed}"
        ws["E2"] = f"Basarisiz: {failed}"
        ws["F2"] = f"Atlanan: {skipped}"
        ws["G2"] = f"Oran: %{rate}"

    def _write_header(self, ws, start_row: int = 1):
        for col_idx, header in enumerate(_HEADERS, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = _THIN_BORDER

    def _write_results(
        self,
        ws,
        start_row: int = 2,
        results: Optional[list[dict]] = None,
    ):
        items = results or self._results
        for idx, result in enumerate(items, 1):
            row = start_row + idx - 1
            values = [
                idx,
                result["test_name"],
                result["status"].upper(),
                result["duration_ms"],
                result["domain"],
                result["error_message"],
                result["timestamp"],
            ]
            status_fill = _STATUS_FILLS.get(result["status"])
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.border = _THIN_BORDER
                if col_idx == 3 and status_fill:
                    cell.fill = status_fill
                    cell.font = Font(bold=True)

    @staticmethod
    def _auto_width(ws):
        for column_cells in ws.columns:
            max_length = 0
            col_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    length = len(str(cell.value or ""))
                    if length > max_length:
                        max_length = length
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 4, 60)
