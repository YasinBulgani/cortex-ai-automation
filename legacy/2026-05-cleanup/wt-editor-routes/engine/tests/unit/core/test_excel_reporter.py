"""
Unit tests for core.excel_reporter — Excel Rapor Ureticisi
"""
import pytest
from pathlib import Path

import sys
sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent.parent))


try:
    from core.excel_reporter import ExcelReporter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


@pytest.mark.skipif(not _HAS_OPENPYXL, reason="openpyxl not installed")
class TestExcelReporter:
    def test_create_reporter(self):
        reporter = ExcelReporter("Test Raporu")
        assert reporter.report_name == "Test Raporu"

    def test_add_result(self):
        reporter = ExcelReporter("Test")
        reporter.add_result("Login Testi", "passed", 1200)
        reporter.add_result("Odeme Testi", "failed", 3500, "Timeout")
        assert len(reporter._results) == 2

    def test_save_creates_file(self, tmp_path):
        reporter = ExcelReporter("Rapor")
        reporter.add_result("Test-1", "passed", 500)
        reporter.add_result("Test-2", "failed", 1000, "Hata mesaji")
        reporter.add_result("Test-3", "skipped", 0)

        filepath = reporter.save(tmp_path)
        assert Path(filepath).exists()
        assert filepath.endswith(".xlsx")

    def test_save_with_multi_domain(self, tmp_path):
        reporter = ExcelReporter("Multi-Domain")
        reporter.add_result("Test-1", "passed", 500, domain="girit")
        reporter.add_result("Test-2", "failed", 800, error_message="Err", domain="plus")
        reporter.add_result("Test-3", "passed", 300, domain="girit")

        filepath = reporter.save(tmp_path)
        assert Path(filepath).exists()

        from openpyxl import load_workbook
        wb = load_workbook(filepath)
        sheet_names = wb.sheetnames
        assert "Test Sonuclari" in sheet_names
        assert "girit" in sheet_names
        assert "plus" in sheet_names

    def test_empty_report(self, tmp_path):
        reporter = ExcelReporter("Bos")
        filepath = reporter.save(tmp_path)
        assert Path(filepath).exists()
